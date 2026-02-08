

import numpy as np
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.widgets import Slider
import tkinter.colorchooser as colorchooser
import json
from tkinter import messagebox
import os
import ctypes
import math
from tkinter import simpledialog, filedialog
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gc
from tksheet import Sheet
from PIL import ImageGrab



'''
1.3 版本更新日志

1. 增加保存配置和读取配置选项，可以将页面中设置好的参数直接保存，避免反复修改.
2. 读取数据和保存数据的功能增加对xlsx和csv文件的支持.
3. 保存cdxml文件现在可以选择保存位置.
4. 增加了读取chemdraw文件风格的功能,现在可以使用“Load CDXML Style'选择自己常用风格的chemdraw文件,读取后续保存的也会是同样的风格。同时保存配置的时候也会被存储到配置信息里.
5. 增加了显示数字和显示文字的选项

1.2 版本更新日志

1. 修复了由于ID相同导致圆形图案出现在线条层下方的bug.
2. 修复了调整页面大小影响显示的bug(现在可以自由修改页面大小和纵横比).
3. 优化了页面调整逻辑, 使其更加合理.
4. 重新调整了底部工具栏代码, 现在所有工具都能正常工作.
5. 当基础曲线 = 0 时, 绘制的形状现在是真正的直线.
6. 在表格区域添加了右键菜单, 右键单击表格可以删除行/列, 在左侧添加行, 在顶部添加列.
7. 添加了"显示网格"选项, 允许设置是否显示网格线.
8. 修复了从包含空行的表格导出数据时导致错误的问题.
9. 添加了"打开表格"按钮以打开表格区域.
10. 如果多个标记位于同一位置, 则只绘制一个.

注意: 所有版本保存的 table_data.json 都是通用的, 可以被任何版本读取.

1.1 版本更新日志

1. 自动扩展页面: 页面现在可以自动扩展, 允许无缝添加内容.
2. 设置标签位置: 用户现在可以自定义标签的位置.
3. 单个标记绘制: 添加了在图上绘制单个标记的功能以进行视觉强调.
4. 字体大小自定义: 用户现在可以调整图上文本元素的字体大小.
5. 字体和标签定位: 添加了设置字体与标签中心位置之间距离的功能.
6. 线条粗细自定义: 用户现在可以调整图上绘制的线条粗细.
7. 键粗细自定义: 结构中键的粗细现在可以自定义.

1.0 版本:
EnergyCurvePlot 是一个用于生成化学反应能量变化曲线的工具.
该程序允许用户可视化整个反应过程中的能量演变,
提供对反应机理和能量分布的洞察. 输出是一个 cdxml 文件,
可以在 Chemdraw 中轻松打开和编辑以进行进一步自定义.

由兰州大学 Yatao Lang 开发
'''

cdxml_header = '''<?xml version="1.0" encoding="UTF-8" ?>
<!DOCTYPE CDXML SYSTEM "http://www.cambridgesoft.com/xml/cdxml.dtd" >
<CDXML
 CreationProgram="ChemDraw 22.0.0.22"
 Name="output.cdxml"
 BoundingBox="259.57 429.79 816.37 719.79"
 WindowPosition="0 0"
 WindowSize="-2147483648 0"
 WindowIsZoomed="yes"
 FractionalWidths="yes"
 InterpretChemically="yes"
 ShowAtomQuery="yes"
 ShowAtomStereo="no"
 ShowAtomEnhancedStereo="yes"
 ShowAtomNumber="no"
 ShowResidueID="no"
 ShowBondQuery="yes"
 ShowBondRxn="yes"
 ShowBondStereo="no"
 ShowTerminalCarbonLabels="no"
 ShowNonTerminalCarbonLabels="no"
 HideImplicitHydrogens="no"
 LabelFont="3"
 LabelSize="10"
 LabelFace="96"
 CaptionFont="3"
 CaptionSize="10"
 HashSpacing="2.50"
 MarginWidth="1.60"
 LineWidth="0.60"
 BoldWidth="2"
 BondLength="14.40"
 BondSpacing="18"
 ChainAngle="120"
 LabelJustification="Auto"
 CaptionJustification="Left"
 AminoAcidTermini="HOH"
 ShowSequenceTermini="yes"
 ShowSequenceBonds="yes"
 ShowSequenceUnlinkedBranches="no"
 ResidueWrapCount="40"
 ResidueBlockCount="10"
 PrintMargins="36 36 36 36"
 MacPrintInfo="0003000001200120000000000B6608A0FF84FF880BE309180367052703FC0002000001200120000000000B6608A0000100000064000000010001010100000001270F000100010000000000000000000000000002001901900000000000600000000000000000000100000000000000000000000000000000"
 ChemPropName=""
 ChemPropFormula="Chemical Formula: "
 ChemPropExactMass="Exact Mass: "
 ChemPropMolWt="Molecular Weight: "
 ChemPropMOverZ="m/z: "
 ChemPropAnalysis="Elemental Analysis: "
 ChemPropBoilingPt="Boiling Point: "
 ChemPropMeltingPt="Melting Point: "
 ChemPropCritTemp="Critical Temp: "
 ChemPropCritPres="Critical Pres: "
 ChemPropCritVol="Critical Vol: "
 ChemPropGibbs="Gibbs Energy: "
 ChemPropLogP="Log P: "
 ChemPropMR="MR: "
 ChemPropHenry="Henry&apos;s Law: "
 ChemPropEForm="Heat of Form: "
 ChemProptPSA="tPSA: "
 ChemPropCLogP="CLogP: "
 ChemPropCMR="CMR: "
 ChemPropLogS="LogS: "
 ChemPropPKa="pKa: "
 ChemPropID=""
 ChemPropFragmentLabel=""
 color="0"
 bgcolor="1"
 RxnAutonumberStart="1"
 RxnAutonumberConditions="no"
 RxnAutonumberStyle="Roman"
 RxnAutonumberFormat="(#)"
>'''

font_xml = '''<fonttable>
<font id="3" charset="iso-8859-1" name="Arial"/>
</fonttable>'''

cdxml_footer = "</page></CDXML>"

shape_z_counter = 100
text_z_counter = 300
line_z_counter = 1
global_id = 1
# 为了避免在生成圆形时使用相同的z值
already_draw_target= []

def get_bezier_curve_points_flat(x, y, adjustment_factor=0.05, width_factor=2, adjustment_factor_2=0.1):
    """
    功能:
        生成平滑的贝塞尔曲线控制点, 用于绘制能量曲线
    参数:
        x: 列表, x坐标点
        y: 列表, y坐标点(能量值)
        adjustment_factor: 浮点, 基于距离变化的调整因子, 默认0.05
        width_factor: 浮点, 曲线宽度因子, 默认2
        adjustment_factor_2: 浮点, 基于能量变化的调整因子, 默认0.1
    返回:
        元组(x_adjusted, y, bezier_curves), 其中bezier_curves是包含控制点的字典列表
    """
    if x == []:
        x_adjusted = ['no_data']
        y = ['no_data']
        bezier_curves = ['no_data']
        return x_adjusted, y, bezier_curves
    # 计算相邻能量点之间的能量变化(绝对值)
    delta_energy = np.abs(np.diff(y))
    delta_distance = np.abs(np.diff(x))

    # 初始化x坐标, 从0开始
    x_adjusted = x

    # 根据能量变化调整x坐标
    width_factor_adjusted = []
    for i in range(len(x)-1):
        width_factor_adjusted.append(width_factor + delta_distance[i]*delta_distance[i]* adjustment_factor + delta_energy[i]* adjustment_factor_2)

    x_adjusted = np.array(x_adjusted)
    y = np.array(y)
    width_factor_adjusted = np.array(width_factor_adjusted)

    bezier_curves = []

    # 为每个区间计算贝塞尔控制点
    for i in range(len(x_adjusted) - 1):
        x0, x1 = x_adjusted[i], x_adjusted[i + 1]
        y0, y1 = y[i], y[i + 1]

        # 6点贝塞尔曲线的控制点: 起点, 终点, 以及每个点的两个控制点
        X0, Y0 = x0, y0  # 起点
        X5, Y5 = x1, y1  # 终点
        X1, Y1 = x0 - width_factor_adjusted[i], y0  # 起点左侧的控制点
        X4, Y4 = x1 + width_factor_adjusted[i], y1  # 终点右侧的控制点
        X2, Y2 = x0 + width_factor_adjusted[i], y0  # 起点右侧的控制点
        X3, Y3 = x1 - width_factor_adjusted[i], y1  # 终点左侧的控制点

        bezier_curves.append({
            'X': [X1, X0, X2, X3, X5, X4],
            'Y': [Y1, Y0, Y2, Y3, Y5, Y4]
        })

    return x_adjusted, y, bezier_curves 

def draw_curve(bezier_curves, connect_type='center', bond_length=10, curves_color='3', line_type='Solid', original_x=None, curve_width=0.6):
    """
    功能:
        根据贝塞尔曲线控制点生成XML格式的曲线元素
    参数:
        bezier_curves: 列表, 包含贝塞尔曲线控制点的字典列表
        connect_type: 字符串, 连接类型('center'或'side'), 默认'center'
        bond_length: 浮点, 键长度, 默认10
        curves_color: 字符串, 曲线颜色代码, 默认'3'
        line_type: 字符串, 线条类型, 默认'Solid'
        original_x: 列表, 原始x坐标, 当connect_type='side'时使用
        curve_width: 浮点, 曲线宽度, 默认0.6
    返回:
        列表, 包含XML格式的曲线字符串
    """
    curves_xml = []
    global global_id
    global line_z_counter
    if connect_type == 'center':
        for curve_id, curve in enumerate(bezier_curves, start=1):
            X = curve['X']
            Y = curve['Y']
            scaled_points = [f"{X[i]:.2f} {Y[i]:.2f}" for i in range(len(X))]  # 翻转Y坐标
            curve_points_str = " ".join(scaled_points)
            curves_xml.append(f'<curve id="{global_id}"\n Z="{line_z_counter}\n" color="{curves_color}"\n LineType="{line_type}"\n LineWidth="{curve_width}"\n CurvePoints="{curve_points_str}"\n />')
            global_id += 1
            line_z_counter += 1
    elif connect_type == 'side':
        for curve_id, curve in enumerate(bezier_curves, start=1):
            X = curve['X']
            Y = curve['Y']

            X[0] = X[0] + bond_length*(2*original_x[curve_id-1]+3)
            X[1] = X[1] + bond_length*(2*original_x[curve_id-1]+3)
            X[2] = X[2] + bond_length*(2*original_x[curve_id-1]+3)

            X[3] = X[3] + bond_length*(2*original_x[curve_id]+1)
            X[4] = X[4] + bond_length*(2*original_x[curve_id]+1)
            X[5] = X[5] + bond_length*(2*original_x[curve_id]+1)

            scaled_points = [f"{X[i] :.2f} {Y[i]:.2f}" for i in range(len(X))]  # 翻转Y坐标
            curve_points_str = " ".join(scaled_points)
            curves_xml.append(f'<curve id="{global_id}"\n Z="{line_z_counter}"\n color="{curves_color}"\n LineType="{line_type}"\n LineWidth="{curve_width}"\n CurvePoints="{curve_points_str}"\n />')
            global_id += 1
            line_z_counter += 1
    return curves_xml

def draw_line(center_x, center_y, linetype, linewidth, bond_length=10, bond_color=4, connect_type='side', original_x=None):
    """
    功能:
        生成连接能量点的直线XML元素
    参数:
        center_x: 列表, 中心点x坐标
        center_y: 列表, 中心点y坐标
        linetype: 字符串, 线条类型
        linewidth: 浮点, 线条宽度
        bond_length: 浮点, 键长度, 默认10
        bond_color: 整数, 线条颜色代码, 默认4
        connect_type: 字符串, 连接类型('center'或'side'), 默认'side'
        original_x: 列表, 原始x坐标, 当connect_type='side'时使用
    返回:
        列表, 包含XML格式的线条字符串
    """
    line_xml = []
    global global_id
    global line_z_counter


    if connect_type == 'center':
        # 转换中心坐标
        scale_center_x = center_x
        scale_center_y = center_y
    elif connect_type == 'side':
        # 转换中心坐标
        scale_center_x = [x + bond_length*2*(original_x[idx]+1) for idx, x in enumerate(center_x)]
        scale_center_y = center_y

    for i in range(len(scale_center_x)-1):
        # 计算边界框坐标
        if connect_type == 'center':
            x1 = scale_center_x[i]
            x2 = scale_center_x[i+1]
            y1 = scale_center_y[i]
            y2 = scale_center_y[i+1]
            bounding_box = f"{x1:.2f}  {y1:.2f} {x2:.2f} {y2:.2f}"
        elif connect_type == 'side':
            x1 = scale_center_x[i]+ bond_length
            x2 = scale_center_x[i+1] - bond_length
            y1 = scale_center_y[i]
            y2 = scale_center_y[i+1]
            bounding_box = f"{x1:.2f}  {y1:.2f} {x2:.2f} {y2:.2f}"

        line = ET.Element('arrow', id=str(global_id), Z=str(line_z_counter), LineType=f'{linetype} Bold')
        global_id += 1
        line_z_counter += 1
        # 设置BoundingBox属性
        line.set("BoundingBox", bounding_box)

        # 设置颜色属性(如果提供)
        if bond_color:
            line.set("color", str(bond_color))

        # 设置3D坐标
        head3d = f"{x2:.2f} {scale_center_y[i+1] :.2f} 0"
        tail3d = f"{x1:.2f} {scale_center_y[i] :.2f} 0"
        center3d = f"{scale_center_x[i] :.2f} {scale_center_y[i]:.2f} 0"
        major_axis_end3d = f"{x2:.2f} {y2:.2f} 0"
        minor_axis_end3d = f"{x1:.2f} {y2:.2f} 0"

        line.set("BoldWidth", str(linewidth))
        line.set("Head3D", head3d)
        line.set("Tail3D", tail3d)
        line.set("Center3D", center3d)
        line.set("MajorAxisEnd3D", major_axis_end3d)
        line.set("MinorAxisEnd3D", minor_axis_end3d)

        # 转换为字符串并添加到列表
        line_xml.append(ET.tostring(line, encoding='unicode', method='xml'))

    return line_xml

def generate_rectangle_xml(center_x, center_y, bond_length=10, bond_color=4, connect_type='side', original_x=None, bond_width=2.0):
    """
    功能:
        生成表示能量点的矩形(短线)XML元素
    参数:
        center_x: 列表, 中心点x坐标
        center_y: 列表, 中心点y坐标
        bond_length: 浮点, 键长度, 默认10
        bond_color: 整数, 颜色代码, 默认4
        connect_type: 字符串, 连接类型('center'或'side'), 默认'side'
        original_x: 列表, 原始x坐标, 当connect_type='side'时使用
        bond_width: 浮点, 键宽度, 默认2.0
    返回:
        列表, 包含XML格式的矩形字符串
    """
    rectangles_xml = []
    global global_id
    global shape_z_counter
    global page_width
    global page_high
    global already_draw_target

    if connect_type == 'center':
        # 转换中心坐标
        scale_center_x = center_x
        scale_center_y = center_y
    elif connect_type == 'side':
        # 转换中心坐标
        scale_center_x = [x + bond_length*2*(original_x[idx]+1) for idx, x in enumerate(center_x)]
        scale_center_y = center_y

    # 自动调整页面宽度和高度
    for x in scale_center_x:
        ad_page_width = math.ceil(abs(x/510))
        if ad_page_width > page_width:
            page_width = ad_page_width
    for y in scale_center_y:
        ad_page_high = math.ceil(abs(y/710))
        if ad_page_high > page_high:
            page_high = ad_page_high


    for i in range(len(scale_center_x)):
        coordinate = (scale_center_x[i], scale_center_y[i])
        if coordinate not in already_draw_target:
            already_draw_target.append(coordinate)
        else:
            continue
        # 计算边界框坐标
        x1 = (scale_center_x[i] - bond_length)
        x2 = (scale_center_x[i] + bond_length)
        y1 = scale_center_y[i]
        y2 = scale_center_y[i]
        bounding_box = f"{x1:.2f} {y1:.2f} {x2:.2f} {y2:.2f}"

        arrow = ET.Element('arrow', id=str(global_id), Z=str(shape_z_counter), LineType="Bold", FillType="None", ArrowheadType="Solid")
        global_id += 1
        shape_z_counter += 1
        # 设置BoundingBox属性
        arrow.set("BoundingBox", bounding_box)

        # 设置颜色属性(如果提供)
        if bond_color:
            arrow.set("color", str(bond_color))

        # 设置3D坐标
        head3d = f"{x2:.2f} {scale_center_y[i] :.2f} 0"
        tail3d = f"{x1:.2f} {scale_center_y[i] :.2f} 0"
        center3d = f"{scale_center_x[i] :.2f} {scale_center_y[i]:.2f} 0"
        major_axis_end3d = f"{x2:.2f} {y2:.2f} 0"
        minor_axis_end3d = f"{x1:.2f} {y2:.2f} 0"

        arrow.set("BoldWidth", str(bond_width))
        arrow.set("Head3D", head3d)
        arrow.set("Tail3D", tail3d)
        arrow.set("Center3D", center3d)
        arrow.set("MajorAxisEnd3D", major_axis_end3d)
        arrow.set("MinorAxisEnd3D", minor_axis_end3d)

        # 转换为字符串并添加到列表
        rectangles_xml.append(ET.tostring(arrow, encoding='unicode', method='xml'))

    return rectangles_xml

def generate_circle_xml(center_x, center_y, radius=5, circle_color=8, connect_type='side', original_x=None):
    """
    功能:
        生成表示能量点的圆形标记XML元素
    参数:
        center_x: 列表, 中心点x坐标
        center_y: 列表, 中心点y坐标
        radius: 浮点, 圆形半径, 默认5
        circle_color: 整数, 圆形颜色代码, 默认8
        connect_type: 字符串, 连接类型('center'或'side'), 默认'side'
        original_x: 列表, 原始x坐标, 当connect_type='side'时使用
    返回:
        列表, 包含XML格式的圆形字符串
    """
    import xml.etree.ElementTree as ET
    global shape_z_counter
    global global_id
    global page_width
    global page_high
    global already_draw_target

    circles_xml = []

    if connect_type == 'center':
        # 转换中心坐标
        scale_center_x = center_x
        scale_center_y = center_y
    elif connect_type == 'side':
        # 转换中心坐标
        scale_center_x = [x + radius *2*(original_x[idx]+1) for idx, x in enumerate(center_x)]
        scale_center_y = center_y

    # 自动调整页面宽度和高度
    for x in scale_center_x:
        ad_page_width = math.ceil(abs(x/500))
        if ad_page_width > page_width:
            page_width = ad_page_width
    for y in scale_center_y:
        ad_page_high = math.ceil(abs(y/710))
        if ad_page_high > page_high:
            page_high = ad_page_high

    for i in range(len(scale_center_x)):
        coordinate = (scale_center_x[i], scale_center_y[i])
        if coordinate not in already_draw_target:
            already_draw_target.append(coordinate)
        else:
            continue
        # 计算边界框坐标
        x1 = (scale_center_x[i] - radius)
        x2 = (scale_center_x[i] + radius)
        y1 = (scale_center_y[i] - radius)
        y2 = (scale_center_y[i] + radius)
        bounding_box = f"{x1:.2f} {y1:.2f} {x2:.2f} {y2:.2f}"

        # 创建根元素
        circle_id = 20 + i  # 示例id, 可以参数化
        graphic = ET.Element('graphic', id=str(global_id), Z=str(shape_z_counter), color=str(circle_color), GraphicType="Oval", OvalType="Circle Filled")
        shape_z_counter += 1
        global_id += 1
        # 设置BoundingBox属性
        graphic.set("BoundingBox", bounding_box)

        # 设置3D坐标
        center3d = f"{scale_center_x[i]:.2f} {scale_center_y[i]:.2f} 0"
        major_axis_end3d = f"{x2:.2f} {scale_center_y[i]:.2f} 0"
        minor_axis_end3d = f"{scale_center_x[i]:.2f} {y2:.2f} 0"

        graphic.set("Center3D", center3d)
        graphic.set("MajorAxisEnd3D", major_axis_end3d)
        graphic.set("MinorAxisEnd3D", minor_axis_end3d)

        # 转换为字符串并添加到列表
        circles_xml.append(ET.tostring(graphic, encoding='unicode', method='xml'))

    return circles_xml

def generate_text_cdxml(center_x, center_y, energy_text_list, target_text_list, location, target_layout, target_location, text_space, text_base_movement, target_move, target_move_y, text_color=4, connect_type='side', original_x=None, font_size=10, font_type=3, Z_value=70, show_numbers=True, show_targets=True):
    """
    功能:
        生成包含能量值和目标文本的CDXML文本元素
    参数:
        center_x: 列表, 每个文本元素的x坐标
        center_y: 列表, 每个文本元素的y坐标
        energy_text_list: 列表, 能量文本列表
        target_text_list: 列表, 目标文本列表
        location: 列表, 文本位置. sc sw sa ss sd cc cw cs ca cd
        target_layout: 字符串, 目标位置的全局设置, 组合位置
        target_location: 字符串, 目标位置的全局设置. c w s a d
        text_space: 浮点, 文本到标记的间距
        text_base_movement: 浮点, 将文本移动到标记中心
        target_move: 浮点, 由目标大小引起的移动(bond_length/radius)
        target_move_y: 浮点, y方向的目标移动
        text_color: 整数, 文本颜色代码, 默认4
        connect_type: 字符串, 连接类型('center'或'side'), 默认'side'
        original_x: 列表, 原始x值, 当connect_type='side'时使用
        font_size: 整数, 字体大小, 默认10
        font_type: 整数, 字体类型, 默认3
        Z_value: 整数, <t>标签的Z值, 默认70
        show_numbers: 布尔, 是否显示能量数字, 默认True
        show_targets: 布尔, 是否显示目标文本, 默认True
    返回:
        列表, 字符串格式的cdxml文本元素列表
    """

    cdxml_elements = []
    def add_text_xml(p_x, p_y, text, text_color, font_size, font_type, Z_value, cdxml_elements, i):
        global global_id
        global text_z_counter
        if text != '':
            text = str(text)

                # if len(text) == 1 and text.isalpha() and text.isupper():
                #     p_x -= 3

            # 创建<t>元素及其对应的属性
            t_element = ET.Element('t', id=str(global_id), p=f"{p_x:.2f} {p_y:.2f}", Z=str(text_z_counter),CaptionJustification="Center",Justification="Center",LineHeight="auto",InterpretChemically="no")
            global_id += 1
            text_z_counter += 1
            # 创建<s>元素用于文本样式
            s_element = ET.SubElement(t_element, 's', font=str(font_type), size=str(font_size), color=str(text_color))
            s_element.text = text

            # 将元素转换为字符串并添加到列表
            cdxml_elements.append(ET.tostring(t_element, encoding='unicode', method='xml'))

    # 根据connect_type调整坐标
    if connect_type == 'center':
        # 不进行转换, 直接使用提供的坐标
        scale_center_x = center_x
        scale_center_y = [y + text_base_movement  for y in center_y]
    elif connect_type == 'side' and original_x is not None:
        # 根据original_x和radius调整x坐标
        scale_center_x = [x +  target_move * 2 *(original_x[idx]+1) for idx, x in enumerate(center_x)]
        scale_center_y = [y + text_base_movement  for y in center_y]
    else:
        raise ValueError("Invalid 'connect_type' or missing 'original_x' for 'side'.")
    
    target_move_y += font_size*0.5

    for idx in range (len(scale_center_x)):
        i = idx
        coordinate = (scale_center_x[i], scale_center_y[i])
        if coordinate not in already_draw_target:
            already_draw_target.append(coordinate)
        else:
            continue
        if location[idx].lower() == 'sc':
            if show_numbers:
                add_text_xml(scale_center_x[idx],scale_center_y[idx]-text_space-target_move_y, energy_text_list[idx], text_color, font_size, font_type, Z_value, cdxml_elements, i)
            if target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'cw':
            if show_numbers:
                if target_text_list[idx] and show_targets:
                    combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                else:
                    combine_text = str(energy_text_list[idx])
                add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
            elif target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'cs':
            if show_numbers:
                if target_text_list[idx] and show_targets:
                    combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                else:
                    combine_text = str(energy_text_list[idx])
                add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
            elif target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'ca':
            if show_numbers:
                if target_text_list[idx] and show_targets:
                    combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                else:
                    combine_text = str(energy_text_list[idx])
                add_text_xml(scale_center_x[idx]-text_space-target_move-len(combine_text)/2*font_size*0.5, scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
            elif target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx]-text_space-target_move-len(target_text_list[idx])/2*font_size*0.5, scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'cd':
            if show_numbers:
                if target_text_list[idx] and show_targets:
                    combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                else:
                    combine_text = str(energy_text_list[idx])
                add_text_xml(scale_center_x[idx]+text_space+target_move+len(combine_text)/2*font_size*0.5, scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
            elif target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx]+text_space+target_move+len(target_text_list[idx])/2*font_size*0.5, scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'cc':
            if show_numbers:
                if target_text_list[idx] and show_targets:
                    combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                else:
                    combine_text = str(energy_text_list[idx])
                add_text_xml(scale_center_x[idx], scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
            elif target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

        elif location[idx].lower() == 'sw':
            if show_numbers:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
            if target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-font_size-target_move_y,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'ss':
            if show_numbers:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
            if target_text_list[idx] and show_targets:
                add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+font_size+target_move_y,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'sa':
            if target_text_list[idx] and show_targets:
                if show_numbers:
                    add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]-font_size/2,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(target_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]+font_size/2,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
            else:
                if show_numbers:
                    add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx],str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
        elif location[idx].lower() == 'sd':
            if target_text_list[idx] and show_targets:
                if show_numbers:
                    add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]-font_size/2,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(target_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]+font_size/2,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
            else:
                if show_numbers:
                    add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx],str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
        else:
            if target_layout == 'sperate':
                if target_location == 'c': #sc
                    if show_numbers:
                        add_text_xml(scale_center_x[idx],scale_center_y[idx]-text_space-target_move_y, energy_text_list[idx], text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    if target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif target_location == 'w': #sw
                    if show_numbers:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    if target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-font_size-target_move_y,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif target_location == 's': #ss
                    if show_numbers:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    if target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+font_size+target_move_y,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif target_location == 'a': #sa
                    if target_text_list[idx] and show_targets:
                        if show_numbers:
                            add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]-font_size/2,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                        add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(target_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]+font_size/2,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    else:
                        if show_numbers:
                            add_text_xml(scale_center_x[idx]-text_space-target_move-len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx],str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                elif target_location == 'd': #sd
                    if target_text_list[idx] and show_targets:
                        if show_numbers:
                            add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]-font_size/2,str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                        add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(target_text_list[idx]))/2*font_size*0.5, scale_center_y[idx]+font_size/2,str(target_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    else:
                        if show_numbers:
                            add_text_xml(scale_center_x[idx]+text_space+target_move+len(str(energy_text_list[idx]))/2*font_size*0.5, scale_center_y[idx],str(energy_text_list[idx]),text_color, font_size, font_type, Z_value, cdxml_elements, i)

            elif target_layout == 'combine':
                if target_location == 'c': #cc
                    if show_numbers:
                        if target_text_list[idx] and show_targets:
                            combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                        else:
                            combine_text = str(energy_text_list[idx])
                        add_text_xml(scale_center_x[idx], scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    elif target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif  target_location == 'w': #cw
                    if show_numbers:
                        if target_text_list[idx] and show_targets:
                            combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                        else:
                            combine_text = str(energy_text_list[idx])
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    elif target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]-text_space-target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif  target_location == 's': #cs
                    if show_numbers:
                        if target_text_list[idx] and show_targets:
                            combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                        else:
                            combine_text = str(energy_text_list[idx])
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    elif target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx], scale_center_y[idx]+text_space+target_move_y,target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif  target_location == 'a': #ca
                    if show_numbers:
                        if target_text_list[idx] and show_targets:
                            combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                        else:
                            combine_text = str(energy_text_list[idx])
                        add_text_xml(scale_center_x[idx]-text_space-target_move-len(combine_text)/2*font_size*0.5, scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    elif target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx]-text_space-target_move-len(target_text_list[idx])/2*font_size*0.5, scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)

                elif  target_location == 'd': #cd
                    if show_numbers:
                        if target_text_list[idx] and show_targets:
                            combine_text = f'{energy_text_list[idx]} {target_text_list[idx]}'
                        else:
                            combine_text = str(energy_text_list[idx])
                        add_text_xml(scale_center_x[idx]+text_space+target_move+len(combine_text)/2*font_size*0.5, scale_center_y[idx],combine_text,text_color, font_size, font_type, Z_value, cdxml_elements, i)
                    elif target_text_list[idx] and show_targets:
                        add_text_xml(scale_center_x[idx]+text_space+target_move+len(target_text_list[idx])/2*font_size*0.5, scale_center_y[idx],target_text_list[idx],text_color, font_size, font_type, Z_value, cdxml_elements, i)
                        
    return cdxml_elements

def save_cdxml_file(cdxml_string, filename="output.cdxml"):
    """
    功能:
        将CDXML字符串保存到文件, 弹出文件保存对话框让用户选择保存位置
    参数:
        cdxml_string: 字符串, CDXML格式的内容
        filename: 字符串, 默认文件名, 默认"output.cdxml"
    返回:
        无
    """
    # 弹出文件保存对话框
    file_path = filedialog.asksaveasfilename(
        defaultextension=".cdxml",
        filetypes=[("CDXML文件", "*.cdxml"), ("所有文件", "*.*")],
        initialfile=filename,
        title="保存CDXML文件"
    )

    # 如果用户取消了保存操作, 则返回
    if not file_path:
        print("保存操作已取消")
        return

    # 保存文件
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(cdxml_string)

    print(f"CDXML文件已成功保存到: {file_path}")

def adjust_line_width_based_on_chart_ratio(ax, scaled_x_adjusted, scaled_y_points):
    """
    通过图表的宽高比例和scaled_x_adjusted, scaled_y_points中最大差值，确定是绘图的坐标轴依据哪个变化的，
    并根据此调整线宽。
    
    参数：
        ax: 当前的matplotlib轴 (ax) 对象
        scaled_x_adjusted: x轴上的调整坐标列表
        scaled_y_points: y轴上的调整坐标列表
        line_width: 默认线宽
    """
    line_width = 30
    
    # 计算scaled_x_adjusted和scaled_y_points中的最大差值
    x_diff = max(scaled_x_adjusted) - min(scaled_x_adjusted)
    y_diff = max(scaled_y_points) - min(scaled_y_points)

    # 获取图表的位置和大小比例 [left, bottom, width, height]
    fig_size = fig.get_size_inches()
    chart_width = fig_size[0]
    chart_height = fig_size[1]


    # 计算图表的比例，比较 x 和 y 轴的变化
    if x_diff/chart_width > y_diff / chart_height:
        # 如果 x 轴的变化范围占比更大，依据 x 轴调整线宽
        ratio = chart_width / x_diff
        adjusted_line_width = line_width * ratio
    else:
        # 如果 y 轴的变化范围占比更大，依据 y 轴调整线宽
        ratio = chart_height / y_diff
        adjusted_line_width = line_width * ratio

    # 返回调整后的线宽
    return adjusted_line_width

hlines = []
curve_lines = []
target_texts = []
energy_texts= []
combine_texxts = []
size_markers = []

def get_contrast_color(hex_color):
    """
    功能:
        根据背景颜色计算对比文字颜色(黑或白), 确保文字可见
    参数:
        hex_color: 十六进制颜色字符串, 如 '#FF0000'
    返回:
        字符串, '#FFFFFF'(白色) 或 '#000000'(黑色)
    """
    # 移除#号
    hex_color = hex_color.lstrip('#')
    # 转换为RGB
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    # 计算亮度
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    return '#FFFFFF' if luminance < 0.5 else '#000000'

def enable_windows_dpi_awareness():
    """
    功能:
        在Windows上启用DPI感知, 避免Tk坐标与截图像素坐标不一致
    参数:
        无
    返回:
        无
    """
    if os.name != 'nt':
        return

    # 优先启用 Per-Monitor V2, 兼容高DPI和多显示器
    try:
        DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2 = -4
        ctypes.windll.user32.SetProcessDpiAwarenessContext(
            ctypes.c_void_p(DPI_AWARENESS_CONTEXT_PER_MONITOR_AWARE_V2)
        )
        return
    except Exception:
        pass

    # 兼容较老Windows版本
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
        return
    except Exception:
        pass

    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

def interactive_bezier_curve():
    """
    功能:
        创建交互式贝塞尔曲线绘图界面, 用于能量曲线的可视化和编辑
    参数:
        无
    返回:
        无
    """
    # 在创建Tk窗口前启用DPI感知, 修复取色偏移
    enable_windows_dpi_awareness()

    # 用于存储从CDXML文件读取的风格信息
    loaded_cdxml_header = None
    loaded_font_xml = None
    loaded_font_id = None  # 存储读取的字体ID
    loaded_font_name = None  # 存储读取的字体名称

    # 取色器状态管理
    picker_state = {
        'target_row': None,
        'target_col': None
    }
    picker_window = None
    preview_frame = None

    def parse_data(data):
        """
        功能:
            解析表格中的数据字符串, 提取能量值、目标文本和位置信息
        参数:
            data: 字符串, 格式为"能量值(目标文本), 位置"
        返回:
            列表[y_value, target, location], 包含解析后的能量值、目标文本和位置
        """
        # 去除两端的空白字符
        data = data.strip()

        # 初始化默认值
        y_value = None
        target = ""
        location = "bs"  # 默认location值为"bs"

        # 替换中文括号和中文逗号为英文括号和逗号
        data = data.replace('（', '(').replace('）', ')').replace('，', ',')


        if ',' in data:
            y_split = data.split(',')
            y_and_target = y_split[0]
            if y_split[1].strip():
                location = y_split[1].strip()
        else:
            y_and_target = data


        if '(' in y_and_target:
            y_value = y_and_target.split('(')[0]
            target = y_and_target.split('(')[1].split(')')[0].strip()

        else:
            y_value = y_and_target


        return [y_value, target, location]

    def draw_benzene(ax, center_x, center_y, line_width, hexagon_side, bond_gap, bond_length_ratio):
        """
        功能:
            在matplotlib轴上绘制类似苯环的六边形, 带有较短的交替双键
        参数:
            ax: matplotlib.axes.Axes对象, 要绘制的轴
            center_x: 浮点, 苯环中心的x坐标
            center_y: 浮点, 苯环中心的y坐标
            line_width: 浮点, 六边形和双键的宽度
            hexagon_side: 浮点, 六边形每条边的长度
            bond_gap: 浮点, 双键线条之间的间隙
            bond_length_ratio: 浮点, 键长度相对于完整边长的比例
        返回:
            无
        """
        # 计算六边形顶点的坐标
        angles = np.linspace(0, 2 * np.pi, 7)  # 6个顶点 + 闭合点
        x = center_x + hexagon_side * np.cos(angles)
        y = center_y + hexagon_side * np.sin(angles)

        # 绘制六边形
        hexagon_line, = ax.plot(x, y, 'k-', linewidth=line_width)  # 提取Line2D对象
        size_markers.append(hexagon_line)

        # 双键索引
        bond_indices = [(0, 1), (2, 3), (4, 5)]  # 双键的顶点对
        for i, j in bond_indices:
            # 计算键的方向向量
            dx = x[j] - x[i]
            dy = y[j] - y[i]
            length = np.sqrt(dx**2 + dy**2)
            dx /= length
            dy /= length

            # 计算较短键的起点和终点
            x_start = x[i] + (1 - bond_length_ratio) / 2 * length * dx
            y_start = y[i] + (1 - bond_length_ratio) / 2 * length * dy
            x_end = x[j] - (1 - bond_length_ratio) / 2 * length * dx
            y_end = y[j] - (1 - bond_length_ratio) / 2 * length * dy

            # 第二条线的垂直偏移
            perp_dx = dy  # 旋转90度
            perp_dy = -dx

            # 绘制较短的双键作为两条靠近的线
            double_bond_line, = ax.plot([x_start - bond_gap * perp_dx, x_end - bond_gap * perp_dx],
                                        [y_start - bond_gap * perp_dy, y_end - bond_gap * perp_dy],
                                        'k-', linewidth=line_width)
            size_markers.append(double_bond_line)

    def update_plot(*args):
        """
        功能:
            更新绘图, 根据用户输入的参数重新绘制能量曲线
        参数:
            *args: 可变参数, 用于接收事件参数
        返回:
            无
        """
        nonlocal loaded_font_name

        try:
            # 获取用户输入参数
            adjustment_factor = float(adjustment_factor_var.get())
            width_factor = float(width_factor_var.get())
            adjustment_factor_2 = float(adjustment_factor_2_var.get())
            scale_factor = float(scale_factor_var.get())
            scale_factor_x = float(scale_factor_x_var.get())
            scale_factor_y = float(scale_factor_y_var.get())
            bond_length = float(bond_length_var.get())
            radius = float(radius_var.get())
            connect_type = connect_type_var.get()
            shape_type = shape_type_var.get()
            line_type = line_type_var.get()
            font_size = font_size_var.get()
            text_space = text_space_var.get()
            show_marker = show_marker_var.get()
            target_layout = target_layout_var.get()
            target_location = target_location_var.get()
            curve_width = curve_width_var.get()
            bond_width = bond_width_var.get()
            grid = grid_var.get()

            # 确定使用的字体名称, 并验证字体是否可用
            if loaded_font_name is not None:
                import matplotlib.font_manager as fm
                available_fonts = [f.name for f in fm.fontManager.ttflist]
                if loaded_font_name in available_fonts:
                    current_font_name = loaded_font_name
                else:
                    current_font_name = 'Arial'
                    print(f"警告: 字体 '{loaded_font_name}' 在系统中未找到, 使用默认字体 Arial")
            else:
                current_font_name = 'Arial'

            if table is None or not table.winfo_exists():
                print('Table window not detected, please click "Open Table"')
                return

            global already_draw_target
            already_draw_target = []
            

            x_min_previous, x_max_previous = ax.get_xlim()
            y_min_previous, y_max_previous = ax.get_ylim()

            x_range_previous = abs(x_max_previous - x_min_previous)
            y_range_previous = abs(y_max_previous - y_min_previous)

            if x_range_previous >10 and y_range_previous >10:
                contain_ax = True
            else:
                contain_ax = False
            
            # 清除之前的绘图
            ax.clear()
            ax.tick_params(axis='both', which='both', bottom=False, top=False,
                left=False, right=False, labelbottom=False, labelleft=False)
            # ax.set_title("Interactive Bezier Curve with Shapes")
            # ax.set_xlabel("Adjusted Coordinate")
            # ax.set_ylabel("Potential Energy")
            if grid:
                ax.grid(True)

            # 反转Y轴
            ax.invert_yaxis()

            # 设置相等的纵横比以保持形状一致性
            ax.set_aspect("equal", adjustable="datalim")

            # line_width = adjust_line_width_based_on_chart_ratio(ax, x_total, y_total)
            line_width = 10
            key_width =  line_width*3.38

            # 从表格获取数据并为每一行创建y列表
            total_rows = table.total_rows()
            for row in range(total_rows):
                # 读取当前行的所有数据
                values = []
                for col in range(table.total_columns()):
                    values.append(table.get_cell_data(row, col))

                # 跳过结构不完整的临时行，避免索引越界
                if len(values) < 3:
                    continue

                original_y = []
                original_x = []
                target = []
                location = []
                for index, value in enumerate(values[3:]):
                    value = str(value)
                    if str(value).strip():  # If value is not empty
                        try:
                            value = parse_data(value)
                            original_y.append(float(value[0]))
                            target.append(value[1])
                            location.append(value[2])
                            original_x.append(index + 1)
                        except ValueError:
                            print(f"Invalid value at row {row}, column {index + 2}. Please correct.")
                            return

                curve_color = values[0]
                marker_color = values[1]
                text_color = values[2]

                # 调用贝塞尔曲线计算函数
                x_adjusted, y_points, bezier_curves = get_bezier_curve_points_flat(
                original_x, original_y, adjustment_factor, width_factor, adjustment_factor_2
                )

                if  x_adjusted[0] == 'no_data': # 无数据行
                    continue

                # 缩放贝塞尔曲线
                for curve in bezier_curves:
                    X = curve['X']
                    Y = curve['Y']
                    for i in range(len(X)):
                        X[i] = X[i] * scale_factor *scale_factor_x* 10 + 50
                        Y[i] = 500 - Y[i] * scale_factor*scale_factor_y / 2

                # 缩放调整后的坐标和势能值
                scaled_x_adjusted = [x * scale_factor*scale_factor_x * 10 + 50 for x in x_adjusted]
                scaled_y_points = [500 - y * scale_factor*scale_factor_y / 2 for y in y_points]

                fontsize = line_width*15 # 随便设置的大小

                # 根据connect_type绘制贝塞尔曲线
                for i, curve in enumerate(bezier_curves):
                    X = curve['X']
                    Y = curve['Y']
                    t = np.linspace(0, 1, 200)

                    if connect_type == "center":
                        # 连续曲线
                        Bx = (1 - t)**3 * X[1] + 3 * (1 - t)**2 * t * X[2] + 3 * (1 - t) * t**2 * X[3] + t**3 * X[4]
                        By = (1 - t)**3 * Y[1] + 3 * (1 - t)**2 * t * Y[2] + 3 * (1 - t) * t**2 * Y[3] + t**3 * Y[4]
                        if line_type == "Dashed":
                            curve_line = ax.plot(Bx, By, color=curve_color,linewidth=line_width,linestyle=line_type.lower(),dashes=(5*0.6/curve_width, 5*0.6/curve_width))
                            curve_lines.append(curve_line)
                        elif line_type == "Solid":
                            curve_line = ax.plot(Bx, By, color=curve_color,linewidth=line_width,linestyle=line_type.lower())
                            curve_lines.append(curve_line)

                    elif connect_type == "side":
                        # 非连续曲线, 调整端点
                        if shape_type == "line":
                            offset = bond_length 
                        elif shape_type == "circle":
                            offset = radius  
                        else:
                            offset = 0

                        X[0] = X[0] + offset*(2*original_x[i]+1)
                        X[1] = X[1] + offset*(2*original_x[i]+1)
                        X[2] = X[2] + offset*(2*original_x[i]+1)

                        X[3] = X[3] + offset*(2*original_x[i+1]-1)
                        X[4] = X[4] + offset*(2*original_x[i+1]-1)
                        X[5] = X[5] + offset*(2*original_x[i+1]-1)

                            
                        Bx = (1 - t)**3 * X[1] + 3 * (1 - t)**2 * t * X[2] + 3 * (1 - t) * t**2 * X[3] + t**3 * X[4]
                        By = (1 - t)**3 * Y[1] + 3 * (1 - t)**2 * t * Y[2] + 3 * (1 - t) * t**2 * Y[3] + t**3 * Y[4]
                        if line_type == "Dashed":
                            curve_line = ax.plot(Bx, By, color=curve_color,linewidth=line_width,linestyle=line_type.lower(),dashes=(5*0.6/curve_width, 5*0.6/curve_width))
                            curve_lines.append(curve_line)
                        elif line_type == "Solid":
                            curve_line = ax.plot(Bx, By, color=curve_color,linewidth=line_width,linestyle=line_type.lower())
                            curve_lines.append(curve_line)

                # 绘制水平线或圆形
                if shape_type == "line":
                    text_shape_space = bond_length
                    text_shape_space_y = bond_width/2 + font_size*0.5
                    if connect_type == "side":
                        for idx, (x, y) in enumerate(zip(scaled_x_adjusted, scaled_y_points)):
                            center_x = x + bond_length*2*(original_x[idx])
                            coordinate = (center_x, y)
                            if coordinate not in already_draw_target:
                                already_draw_target.append(coordinate)
                            else:
                                continue
                            print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space, text_shape_space_y,idx, y, center_x,target_layout,target_location, current_font_name)
                            hline = ax.hlines(y, center_x - bond_length, center_x + bond_length, colors=marker_color, linewidth=key_width,zorder=100)
                            hlines.append(hline)
                            
                    elif connect_type == "center":
                        for idx, (x, y) in enumerate(zip(scaled_x_adjusted, scaled_y_points)):
                            coordinate = (x, y)
                            if coordinate not in already_draw_target:
                                already_draw_target.append(coordinate)
                            else:
                                continue
                            print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space,text_shape_space_y, idx, y, x,target_layout,target_location, current_font_name)
                            hline = ax.hlines(y, x - bond_length, x + bond_length, colors=marker_color, linewidth=key_width,zorder=100)
                            hlines.append(hline)
                            
    
                elif shape_type == "circle":
                    text_shape_space = radius
                    text_shape_space_y = radius + font_size*0.5
                    if connect_type == "side":
                        for idx, (x, y) in enumerate(zip(scaled_x_adjusted, scaled_y_points)):
                            center_x = x + radius*2*(original_x[idx])
                            coordinate = (center_x, y)
                            if coordinate not in already_draw_target:
                                already_draw_target.append(coordinate)
                            else:
                                continue
                            print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space, text_shape_space_y,idx, y, center_x,target_layout,target_location, current_font_name)
                            circle = plt.Circle((center_x, y), radius, color=marker_color, fill=True,zorder=100)
                            ax.add_artist(circle)
                            
                    elif connect_type == "center":
                        for idx, (x, y) in enumerate(zip(scaled_x_adjusted, scaled_y_points)):
                            coordinate = (x, y)
                            if coordinate not in already_draw_target:
                                already_draw_target.append(coordinate)
                            else:
                                continue
                            print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space, text_shape_space_y,idx, y, x,target_layout,target_location)
                            circle = plt.Circle((x, y), radius, color=marker_color, fill=True,zorder=100)
                            ax.add_artist(circle)
                            


                elif shape_type == "None":
                    text_shape_space = 0
                    text_shape_space_y = 0 + font_size*0.5
                    for idx, (x, y) in enumerate(zip(scaled_x_adjusted, scaled_y_points)):
                        coordinate = (x, y)
                        if coordinate not in already_draw_target:
                            already_draw_target.append(coordinate)
                        else:
                            continue
                        print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space,text_shape_space_y, idx, y, x,target_layout,target_location, current_font_name)


            
            canvas.draw()

            if contain_ax == True:
                ax.set_xlim(x_min_previous, x_max_previous)
                ax.set_ylim(y_min_previous, y_max_previous)

            x_min, x_max = ax.get_xlim()
            y_min, y_max = ax.get_ylim()
            center_x = (x_min+x_max)/2
            center_y = (y_min+y_max)/2

            # 缩放因子计算
            x_range = x_max - x_min
            y_range = y_min - y_max
            width, height = canvas.get_width_height()

            fontsize = font_size/max(x_range/width, y_range/height)*0.7
            hline_linewidth = bond_width/max(x_range/width, y_range/height)*0.7
            curve_linewidth = curve_width/max(x_range/width, y_range/height)*0.7
            size_marker_linewidth = 0.6/max(x_range/width, y_range/height)*0.7


            if contain_ax == True:
                ax.set_xlim(x_min_previous, x_max_previous)
                ax.set_ylim(y_min_previous, y_max_previous)

            for text_obj in ax.texts:
                # current_fontsize = text_obj.get_fontsize()
                # print(current_fontsize)
                text_obj.set_fontsize(fontsize)

            for hline_obj in hlines:
                hline_obj.set_linewidth(hline_linewidth)

            for curve_obj in curve_lines:
                for curve in curve_obj:
                    curve.set_linewidth(curve_linewidth)
            
            #draw maker
            if show_marker:
                draw_benzene(ax, center_x, center_y, size_marker_linewidth,hexagon_side=14, bond_gap=2.5, bond_length_ratio=0.7)

            # 重新绘制画布
            canvas.draw()

                
        except ValueError:
            tk.messagebox.showerror("Input Error", "Please enter valid numeric values.")

    def print_text(font_size, text_space, original_y, target, location, text_color, fontsize, text_shape_space,text_shape_space_y, idx, y, center_x,target_layout,target_location, font_name='Arial'):
        show_numbers = show_numbers_var.get()
        show_targets = show_targets_var.get()

        if location[idx].lower() == 'sc':
            if show_numbers:
                energy_text = ax.text(center_x, y-text_space-text_shape_space_y, str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            if target[idx] and show_targets:
                target_text = ax.text(center_x, y+text_space+text_shape_space_y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'cw':
            if show_numbers:
                if target[idx] and show_targets:
                    combine_text = f'{original_y[idx]} {target[idx]}'
                else:
                    combine_text = str(original_y[idx])
                energy_text = ax.text(center_x, y-text_space-text_shape_space_y, combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            elif target[idx] and show_targets:
                energy_text = ax.text(center_x, y-text_space-text_shape_space_y, str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'cs':
            if show_numbers:
                if target[idx] and show_targets:
                    combine_text = f'{original_y[idx]} {target[idx]}'
                else:
                    combine_text = str(original_y[idx])
                energy_text = ax.text(center_x, y+text_space+text_shape_space_y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            elif target[idx] and show_targets:
                energy_text = ax.text(center_x, y+text_space+text_shape_space_y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'ca':
            if show_numbers:
                if target[idx] and show_targets:
                    combine_text = f'{original_y[idx]} {target[idx]}'
                else:
                    combine_text = str(original_y[idx])
                energy_text = ax.text(center_x-text_space-text_shape_space-len(combine_text)/2*font_size*0.5, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            elif target[idx] and show_targets:
                energy_text = ax.text(center_x-text_space-text_shape_space-len(str(target[idx]))/2*font_size*0.5, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'cd':
            if show_numbers:
                if target[idx] and show_targets:
                    combine_text = f'{original_y[idx]} {target[idx]}'
                else:
                    combine_text = str(original_y[idx])
                energy_text = ax.text(center_x+text_space+text_shape_space+len(combine_text)/2*font_size*0.5, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            elif target[idx] and show_targets:
                energy_text = ax.text(center_x+text_space+text_shape_space+len(str(target[idx]))/2*font_size*0.5, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'cc':
            if show_numbers:
                if target[idx] and show_targets:
                    combine_text = f'{original_y[idx]} {target[idx]}'
                else:
                    combine_text = str(original_y[idx])
                energy_text = ax.text(center_x, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            elif target[idx] and show_targets:
                energy_text = ax.text(center_x, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})  

        elif location[idx].lower() == 'sw':
            if show_numbers:
                energy_text = ax.text(center_x, y-text_space-text_shape_space_y, str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            if target[idx] and show_targets:
                target_text = ax.text(center_x, y-text_space-font_size-text_shape_space_y, str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'ss':
            if show_numbers:
                energy_text = ax.text(center_x, y+text_space+text_shape_space_y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            if target[idx] and show_targets:
                target_text = ax.text(center_x, y+text_space+font_size+text_shape_space_y, str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'sa':
            if target[idx] and show_targets:
                if show_numbers:
                    energy_text = ax.text(center_x-text_space-text_shape_space-len(str(original_y[idx]))/2*font_size*0.5, y-font_size/2 , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                target_text = ax.text(center_x-text_space-text_shape_space-len(str(target[idx]))/2*font_size*0.5, y+font_size/2 , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            else:
                if show_numbers:
                    energy_text = ax.text(center_x-text_space-text_shape_space-len(str(original_y[idx]))/2*font_size*0.5, y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        elif location[idx].lower() == 'sd':
            if target[idx] and show_targets:
                if show_numbers:
                    energy_text = ax.text(center_x+text_space+text_shape_space+len(str(original_y[idx]))/2*font_size*0.5, y-font_size/2 , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                target_text = ax.text(center_x+text_space+text_shape_space+len(str(target[idx]))/2*font_size*0.5, y+font_size/2 , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
            else:
                if show_numbers:
                    energy_text = ax.text(center_x+text_space+text_shape_space+len(str(original_y[idx]))/2*font_size*0.5, y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
        
        else:
            if target_layout == 'sperate':
                if target_location == 'c': #sc
                    if show_numbers:
                        energy_text = ax.text(center_x, y-text_space-text_shape_space_y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                        energy_texts.append(energy_text)
                    if target[idx] and show_targets:
                        target_text = ax.text(center_x, y+text_space+text_shape_space_y, str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                        target_texts.append(target_text)
                elif target_location == 'w': #sw
                    if show_numbers:
                        energy_text = ax.text(center_x, y-text_space-text_shape_space_y  , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    if target[idx] and show_targets:
                        target_text = ax.text(center_x, y-text_space-font_size-text_shape_space_y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                elif target_location == 's': #ss
                    if show_numbers:
                        energy_text = ax.text(center_x, y+text_space+text_shape_space_y  , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    if target[idx] and show_targets:
                        target_text = ax.text(center_x, y+text_space+font_size+text_shape_space_y  , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                elif target_location == 'a': #sa
                    if target[idx] and show_targets:
                        if show_numbers:
                            energy_text = ax.text(center_x-text_space-text_shape_space-len(str(original_y[idx]))/2*font_size*0.5, y-font_size/2 , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                        target_text = ax.text(center_x-text_space-text_shape_space-len(str(target[idx]))/2*font_size*0.5, y+font_size/2 , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    else:
                        if show_numbers:
                            energy_text = ax.text(center_x-text_space-text_shape_space-len(str(original_y[idx]))/2*font_size*0.5, y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                elif target_location == 'd': #sd
                    if target[idx] and show_targets:
                        if show_numbers:
                            energy_text = ax.text(center_x+text_space+text_shape_space+len(str(original_y[idx]))/2*font_size*0.5, y-font_size/2 , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                        target_text = ax.text(center_x+text_space+text_shape_space+len(str(target[idx]))/2*font_size*0.5, y+font_size/2 , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    else:
                        if show_numbers:
                            energy_text = ax.text(center_x+text_space+text_shape_space+len(str(original_y[idx]))/2*font_size*0.5, y , str(original_y[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})

            elif target_layout == 'combine':
                if target_location == 'c': #cc
                    if show_numbers:
                        if target[idx] and show_targets:
                            combine_text = f'{original_y[idx]} {target[idx]}'
                        else:
                            combine_text = str(original_y[idx])
                        energy_text = ax.text(center_x, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    elif target[idx] and show_targets:
                        energy_text = ax.text(center_x, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})

                elif  target_location == 'w': #cw
                    if show_numbers:
                        if target[idx] and show_targets:
                            combine_text = f'{original_y[idx]} {target[idx]}'
                        else:
                            combine_text = str(original_y[idx])
                        energy_text = ax.text(center_x, y-text_space-text_shape_space_y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    elif target[idx] and show_targets:
                        energy_text = ax.text(center_x, y-text_space-text_shape_space_y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})

                elif  target_location == 's': #cs
                    if show_numbers:
                        if target[idx] and show_targets:
                            combine_text = f'{original_y[idx]} {target[idx]}'
                        else:
                            combine_text = str(original_y[idx])
                        energy_text = ax.text(center_x, y+text_space+text_shape_space_y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    elif target[idx] and show_targets:
                        energy_text = ax.text(center_x, y+text_space+text_shape_space_y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                elif  target_location == 'a': #ca
                    if show_numbers:
                        if target[idx] and show_targets:
                            combine_text = f'{original_y[idx]} {target[idx]}'
                        else:
                            combine_text = str(original_y[idx])
                        energy_text = ax.text(center_x-text_space-text_shape_space-len(combine_text)/2*font_size*0.5, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    elif target[idx] and show_targets:
                        energy_text = ax.text(center_x-text_space-text_shape_space-len(str(target[idx]))/2*font_size*0.5, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})

                elif  target_location == 'd': #cd
                    if show_numbers:
                        if target[idx] and show_targets:
                            combine_text = f'{original_y[idx]} {target[idx]}'
                        else:
                            combine_text = str(original_y[idx])
                        energy_text = ax.text(center_x+text_space+text_shape_space+len(combine_text)/2*font_size*0.5, y , combine_text, color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})
                    elif target[idx] and show_targets:
                        energy_text = ax.text(center_x+text_space+text_shape_space+len(str(target[idx]))/2*font_size*0.5, y , str(target[idx]), color=text_color,fontsize=fontsize, ha='center', va='center_baseline',fontdict={'family': font_name})

    class CustomNavigationToolbar(NavigationToolbar2Tk):
        def __init__(self, canvas, parent):
            super().__init__(canvas, parent)
            self.ax = ax
            self.show_all = show_all
            self.auto_zoom = auto_zoom
            self.zoom_mode_active = False  # 标记是否处于缩放模式
            self._zoom_timer = None  # 定时器ID
            self._is_dragging = False  # 标记是否正在右键拖动

            # 连接 Matplotlib 的 button_release_event
            canvas.mpl_connect("button_release_event", self.on_button_release)
            canvas.mpl_connect("button_press_event", self.on_button_press)

        def home(self, *args):
            """扩展'Home'按钮行为"""
            # 调用父类的home方法，执行原本的Home行为
            super().home(*args)
            
            # 执行自定义的show_all函数
            self.show_all()

        def forward(self, *args):
            """重写'Forward'按钮行为"""
            # 调用父类的forward方法，执行原本的Forward行为
            super().forward(*args)

            # 执行自定义的auto_zoom函数
            self.auto_zoom()

        def back(self, *args):
            """重写'Back'按钮行为"""
            # 调用父类的back方法，执行原本的Back行为
            super().back(*args)

            # 执行自定义的auto_zoom函数
            self.auto_zoom()

        def zoom(self, *args):
            """重写'Zoom to Rectangle'按钮行为"""
            super().zoom(*args)  # 调用默认的放大镜功能
            self.zoom_mode_active = True  # 标记进入缩放模式

        def on_button_press(self, event):
            """鼠标按下事件处理函数"""
            if event.button == 3:  # 右键按下
                self._is_dragging = True  # 标记开始拖动
                self._schedule_zoom()  # 启动定时器，每隔0.1秒执行auto_zoom()

        def on_button_release(self, event):
            self.auto_zoom()
            """鼠标释放事件处理函数"""
            if self._is_dragging:  # 右键释放
                self._is_dragging = False  # 标记停止拖动
                self._cancel_zoom()  # 停止定时器
                self.auto_zoom()  # 释放时调用一次auto_zoom()

        def _schedule_zoom(self):
            """每隔0.1秒调用一次auto_zoom()"""
            if self._zoom_timer is None:
                self._zoom_timer = self.canvas.get_tk_widget().after(10, self._zoom_callback)

        def _zoom_callback(self):
            """定时器回调函数"""
            self.auto_zoom()  # 执行自动缩放
            if self._is_dragging:  # 如果拖动还在进行，继续调用定时器
                self._zoom_timer = self.canvas.get_tk_widget().after(10, self._zoom_callback)

        def _cancel_zoom(self):
            """停止定时器"""
            if self._zoom_timer is not None:
                self.canvas.get_tk_widget().after_cancel(self._zoom_timer)
                self._zoom_timer = None

        def configure_subplots(self):
            """重写配置按钮行为"""
            tool = super().configure_subplots()  # 调用父类的方法，显示默认配置工具

            # 遍历控件，绑定滑块事件
            for attr in dir(tool):
                obj = getattr(tool, attr)
                if isinstance(obj, Slider):  # 检查是否是滑块
                    obj.on_changed(self.on_slider_change)

        def on_slider_change(self, val):
            """滑块改变时触发的回调函数"""
            self.auto_zoom()

        def save_figure(self):
            """重写保存按钮行为"""

            # 弹出对话框, 获取用户输入的DPI值
            dpi = simpledialog.askinteger("Save Figure", "Enter DPI (e.g., 100, 200, 300):", minvalue=50, maxvalue=1000)
            if dpi is None:
                print("User cancelled save")
                return

            # 弹出文件保存对话框
            file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                    filetypes=[("PNG files", "*.png"),
                                                            ("JPEG files", "*.jpg"),
                                                            ("All files", "*.*")])
            if not file_path:
                print("User cancelled save")
                return

            # 保存图像
            self.canvas.figure.savefig(file_path, dpi=dpi)
            print(f"Image saved to {file_path}, DPI: {dpi}")

    def show_all():
        ax.set_xlim(-1, 1)
        ax.set_ylim(-1, 1)
        update_plot()

    def auto_zoom():
        font_size = font_size_var.get()
        curve_width = curve_width_var.get()
        bond_width = bond_width_var.get()

        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        # 获得X轴方向和Y轴方向的像素数
        width, height = canvas.get_width_height()

        # 缩放因子计算
        x_range = x_max - x_min
        y_range = y_min - y_max
        
        # 控制页面物理尺寸的精华部分
        fontsize = font_size/max(x_range/width, y_range/height)*0.7
        hline_linewidth = bond_width/max(x_range/width, y_range/height)*0.7
        curve_linewidth = curve_width/max(x_range/width, y_range/height)*0.7
        size_marker_linewidth = 0.6/max(x_range/width, y_range/height)*0.7


        for text_obj in ax.texts:
            # current_fontsize = text_obj.get_fontsize()
            # print(current_fontsize)
            text_obj.set_fontsize(fontsize)

        for hline_obj in hlines:
            hline_obj.set_linewidth(hline_linewidth)

        for curve_obj in curve_lines:
            for curve in curve_obj:
                curve.set_linewidth(curve_linewidth)

        for size_marker_obj in size_markers:
            size_marker_obj.set_linewidth(size_marker_linewidth)

        # 重新绘制画布
        canvas.draw()

    def on_resize(event, ax, canvas):
        """在窗口大小变化时触发"""
        auto_zoom()

    def toggle_grid():
        # 切换网格显示
        ax.grid(grid_var.get())
        canvas.draw()
        
    # 创建主窗口
    root = tk.Tk()
    root.title("EnergyCurvePlot v1.3")
    style = ttk.Style(root)
    style.theme_use("vista")  # 使用'vista'主题
    main_frame = tk.Frame(root)
    main_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)



    # 添加Matplotlib图形
    global fig
    fig = Figure(figsize=(8, 6),dpi=100)
    ax = fig.add_subplot(111)
    # ax.set_title("Interactive Bezier Curve")
    # ax.set_xlabel("Adjusted Coordinate")
    # ax.set_ylabel("Potential Energy")
    ax.tick_params(axis='both', which='both', bottom=False, top=False,
               left=False, right=False, labelbottom=False, labelleft=False)
    ax.grid(True)
    canvas = FigureCanvasTkAgg(fig, master=main_frame)
    toolbar = CustomNavigationToolbar(canvas, main_frame)

    toolbar.update()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    canvas.mpl_connect("resize_event", lambda event: on_resize(event, ax, canvas))
    fig.tight_layout() 

    
    def on_zoom(event):
        """
        根据缩放动态调整文本(ax.text)的字体大小，同时调整坐标范围。
        支持放大和缩小
        """
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()
        font_size = font_size_var.get()
        curve_width = curve_width_var.get()
        bond_width = bond_width_var.get()

        # 缩放因子计算
        x_range = x_max - x_min
        y_range = y_max - y_min

        # 根据鼠标滚轮事件确定缩放方向
        if event.button == 'up':  # 放大
            x_min_new = x_min + x_range * 0.1
            x_max_new = x_max - x_range * 0.1
            y_min_new = y_min + y_range * 0.1
            y_max_new = y_max - y_range * 0.1
        elif event.button == 'down':  # 缩小
            x_min_new = x_min - x_range * 0.1
            x_max_new = x_max + x_range * 0.1
            y_min_new = y_min - y_range * 0.1
            y_max_new = y_max + y_range * 0.1

        else:
            return  # 非缩放事件直接返回
        
        # 更新坐标范围
        ax.set_xlim(x_min_new, x_max_new)
        ax.set_ylim(y_min_new, y_max_new)

        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()

        # 获得X轴方向和Y轴方向的像素数
        width, height = canvas.get_width_height()

        # 缩放因子计算
        x_range = x_max - x_min
        y_range = y_min - y_max
        
        # 控制页面物理尺寸的精华部分
        fontsize = font_size/max(x_range/width, y_range/height)*0.7
        hline_linewidth = bond_width/max(x_range/width, y_range/height)*0.7
        curve_linewidth = curve_width/max(x_range/width, y_range/height)*0.7
        size_marker_linewidth = 0.6/max(x_range/width, y_range/height)*0.7


        for text_obj in ax.texts:
            # current_fontsize = text_obj.get_fontsize()
            # print(current_fontsize)
            text_obj.set_fontsize(fontsize)

        for hline_obj in hlines:
            hline_obj.set_linewidth(hline_linewidth)

        for curve_obj in curve_lines:
            for curve in curve_obj:
                curve.set_linewidth(curve_linewidth)

        for size_marker_obj in size_markers:
            size_marker_obj.set_linewidth(size_marker_linewidth)

        # 重新绘制画布
        canvas.draw()
    
    class RightClickMenu:
        """
        功能:
            处理表格的右键菜单, 支持添加/删除行和列
        """
        def __init__(self, table, start_color_picker_func):
            self.table = table
            self.start_color_picker_func = start_color_picker_func
            # 禁用tksheet默认右键菜单
            self.table.disable_bindings("right_click_popup_menu")
            # 绑定自定义右键菜单
            self.table.bind("<Button-3>", self.show_context_menu)

            # 创建上下文菜单
            self.menu = tk.Menu(root, tearoff=0)
            self.menu.add_command(label="删除行", command=self.delete_row)
            self.menu.add_command(label="删除列", command=self.delete_column)
            self.menu.add_command(label="在上方添加行", command=self.add_row)
            self.menu.add_command(label="在下方添加行", command=self.add_row_below)
            self.menu.add_command(label="在左侧添加列", command=self.add_column)
            self.menu.add_command(label="在右侧添加列", command=self.add_column_right)

            self.selected_row = None
            self.selected_col = None

        def pick_screen_color(self):
            """
            功能:
                从右键菜单触发屏幕取色
            参数:
                无
            返回:
                无
            """
            # 设置选中状态
            self.table.select_cell(self.selected_row, self.selected_col)
            # 调用取色函数
            self.start_color_picker_func()

        def show_context_menu(self, event):
            """
            功能:
                显示右键菜单, 根据列类型显示不同选项
            参数:
                event: 鼠标事件对象
            返回:
                无
            """
            # 获取当前选中的单元格
            selected = self.table.get_currently_selected()
            if selected:
                self.selected_row = selected.row
                self.selected_col = selected.column

                # 清空菜单
                self.menu.delete(0, tk.END)

                # 颜色列(0-2列)显示取色选项
                if self.selected_col in [0, 1, 2]:
                    self.menu.add_command(label="屏幕取色", command=self.pick_screen_color)
                    self.menu.add_separator()

                # 数据列(第3列及之后)显示原有选项
                if self.selected_col >= 3:
                    self.menu.add_command(label="删除行", command=self.delete_row)
                    self.menu.add_command(label="删除列", command=self.delete_column)
                    self.menu.add_command(label="在上方添加行", command=self.add_row)
                    self.menu.add_command(label="在下方添加行", command=self.add_row_below)
                    self.menu.add_command(label="在左侧添加列", command=self.add_column)
                    self.menu.add_command(label="在右侧添加列", command=self.add_column_right)

                self.menu.post(event.x_root, event.y_root)

        def delete_row(self):
            """删除选中的行"""
            if self.selected_row is not None:
                self.table.delete_rows([self.selected_row])
                self.selected_row = None

        def delete_column(self):
            """删除选中的列, 保留至少前3列"""
            if self.selected_col is not None:
                total_cols = self.table.total_columns()
                # 确保至少保留3列(颜色列)
                if total_cols <= 3:
                    messagebox.showwarning("警告", "必须保留至少3列")
                    return

                # 删除列
                self.table.delete_columns([self.selected_col])

                # 更新表头(重新编号E列)
                self._update_column_headers()
                self.selected_col = None

        def add_row(self):
            """在选中行上方添加新行"""
            if self.selected_row is not None:
                num_cols = self.table.total_columns()
                new_values = ['#000000', '#000000', '#000000'] + [''] * (num_cols - 3)
                self.table.insert_row(row=new_values, idx=self.selected_row)

                # 设置新行前3列的背景颜色
                for col_idx in range(3):
                    self.table.highlight_cells(
                        row=self.selected_row,
                        column=col_idx,
                        bg='#000000',
                        fg='#FFFFFF'
                    )

                # 取消新添加行的选中状态
                self.table.deselect("all")

        def add_row_below(self):
            """在选中行下方添加新行"""
            if self.selected_row is not None:
                num_cols = self.table.total_columns()
                new_values = ['#000000', '#000000', '#000000'] + [''] * (num_cols - 3)
                self.table.insert_row(row=new_values, idx=self.selected_row + 1)

                # 设置新行前3列的背景颜色
                for col_idx in range(3):
                    self.table.highlight_cells(
                        row=self.selected_row + 1,
                        column=col_idx,
                        bg='#000000',
                        fg='#FFFFFF'
                    )

                # 取消新添加行的选中状态
                self.table.deselect("all")

        def add_column(self):
            """在选中列左侧添加新列"""
            if self.selected_col is not None and self.selected_col >= 3:
                self._insert_column(self.selected_col)

        def add_column_right(self):
            """在选中列右侧添加新列"""
            if self.selected_col is not None and self.selected_col >= 3:
                self._insert_column(self.selected_col + 1)

        def _insert_column(self, col_idx):
            """
            功能:
                在指定位置插入新列
            参数:
                col_idx: 列索引
            返回:
                无
            """
            # 插入空列
            self.table.insert_column(idx=col_idx, width=100)

            # 为所有行在新列填充空值
            total_rows = self.table.total_rows()
            for row_idx in range(total_rows):
                self.table.set_cell_data(row_idx, col_idx, '')

            # 更新表头
            self._update_column_headers()

        def _update_column_headers(self):
            """
            功能:
                更新表头, 重新编号能量列(E1, E2, E3...)
            参数:
                无
            返回:
                无
            """
            total_cols = self.table.total_columns()
            new_headers = ['Curve Color', 'Marker Color', 'Text Color'] + \
                         [f'E{i+1}' for i in range(total_cols - 3)]
            self.table.headers(new_headers)

    # 绑定放大和缩小事件处理函数到 Matplotlib 图表上
    fig.canvas.mpl_connect('scroll_event', on_zoom)

    def create_table_window():

        def on_color_cell_double_click(event):
            """
            功能:
                处理双击事件, 颜色列弹出颜色选择器, 其他列打开编辑器
            参数:
                event: 鼠标事件对象
            返回:
                无
            """
            # 获取点击的单元格
            selected = table.get_currently_selected()
            if not selected:
                return

            row = selected.row
            col = selected.column

            # 处理前3列(颜色列)
            if col in [0, 1, 2]:
                current_value = table.get_cell_data(row, col)

                # 弹出颜色选择器
                default_color = current_value if current_value and current_value.startswith('#') else '#000000'
                color_code = colorchooser.askcolor(initialcolor=default_color, title="选择颜色")[1]

                if color_code:
                    # 更新单元格值(显示十六进制文本)
                    table.set_cell_data(row, col, color_code)
                    # 更新单元格背景颜色和前景文字颜色
                    table.highlight_cells(
                        row=row,
                        column=col,
                        bg=color_code,
                        fg=get_contrast_color(color_code)
                    )
                    update_plot()
        # 创建一个独立的表格窗口
        table_window = tk.Toplevel(root)
        table_window.resizable(True, True)
        table_window.title("Energy Data Table")

        table_window.geometry("900x350")  # 增加窗口大小以显示所有按钮

        # 创建表格的容器框架
        table_frame = tk.Frame(table_window, width=880, height=250)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(5, 0))
        buttons_frame = tk.Frame(table_window)
        buttons_frame.pack(fill=tk.X, padx=5, pady=5)

        # 创建tksheet表格
        table = Sheet(
            table_frame,
            headers=['Curve Color', 'Marker Color', 'Text Color', 'E1', 'E2', 'E3'],
            height=250,
            width=880,
            show_row_index=True,  # 显示行号
            show_header=True,     # 显示表头
            show_top_left=True,   # 显示左上角空白区域
            empty_horizontal=0,   # 不显示额外的水平空白列
            empty_vertical=0      # 不显示额外的垂直空白行
        )

        # 启用基础绑定
        table.enable_bindings(
            "single_select",      # 单击选择
            "drag_select",        # 拖拽选择
            "row_select",         # 行选择
            "column_select",      # 列选择
            "edit_cell"           # 启用单元格编辑（单击后可直接键盘输入）
        )

        # 颜色列通过颜色选择器修改，避免进入文本编辑器
        table.readonly_columns(columns=[0, 1, 2], readonly=True)

        # 设置列宽
        for col_idx in range(6):  # 初始6列
            table.column_width(column=col_idx, width=120)

        # 设置表格所有显示内容居中对齐（数据区、表头、行号）
        table.table_align(align="center")
        table.header_align(align="center")
        table.row_index_align(align="center")

        table.pack(fill=tk.BOTH, expand=True)

        # 初始化一行数据
        table.insert_row(row=['#000000', '#000000', '#000000', 0.0, 0.0, 0.0], idx=0)

        # 设置前3列的背景颜色为其颜色值
        for col_idx in range(3):
            color_value = table.get_cell_data(0, col_idx)
            table.highlight_cells(row=0, column=col_idx, bg=color_value, fg='#FFFFFF')

        # 取消初始行的选中状态
        table.deselect("all")

        # 绑定编辑验证
        table.bind("<<SheetModified>>", lambda e: update_plot())

        # 绑定双击事件, 颜色列打开颜色选择器, 其他列打开编辑器
        table.bind("<Double-Button-1>", on_color_cell_double_click)

        def start_color_picker():
            """
            功能:
                启动屏幕取色器, 验证选择并初始化取色流程
            参数:
                无
            返回:
                无
            """
            nonlocal picker_state, picker_window

            # 获取当前选中的单元格
            selected = table.get_currently_selected()
            if not selected:
                messagebox.showwarning("警告", "请先选择一个颜色单元格")
                return

            row = selected.row
            col = selected.column

            # 验证是否为颜色列(前3列)
            if col not in [0, 1, 2]:
                messagebox.showwarning("警告", "请选择颜色列(Curve Color, Marker Color, Text Color)")
                return

            # 保存选中的单元格位置
            picker_state['target_row'] = row
            picker_state['target_col'] = col

            # 最小化所有窗口
            root.iconify()
            if table_window and table_window.winfo_exists():
                table_window.iconify()

            # 创建全屏取色窗口
            create_picker_window()

        def build_capture_context(screenshot):
            """
            功能:
                构建屏幕坐标到截图像素坐标的映射上下文
            参数:
                screenshot: PIL截图对象
            返回:
                字典, 包含left/top和缩放比例
            """
            img_width, img_height = screenshot.size

            # 默认值(单屏或无法读取系统虚拟屏信息时)
            left = 0
            top = 0
            virtual_width = max(root.winfo_screenwidth(), 1)
            virtual_height = max(root.winfo_screenheight(), 1)

            # Windows下优先使用虚拟桌面坐标, 兼容多显示器
            if os.name == 'nt':
                try:
                    user32 = ctypes.windll.user32
                    left = user32.GetSystemMetrics(76)          # SM_XVIRTUALSCREEN
                    top = user32.GetSystemMetrics(77)           # SM_YVIRTUALSCREEN
                    virtual_width = user32.GetSystemMetrics(78) # SM_CXVIRTUALSCREEN
                    virtual_height = user32.GetSystemMetrics(79)# SM_CYVIRTUALSCREEN
                except Exception:
                    pass

            scale_x = img_width / max(virtual_width, 1)
            scale_y = img_height / max(virtual_height, 1)

            return {
                'left': left,
                'top': top,
                'virtual_width': virtual_width,
                'virtual_height': virtual_height,
                'scale_x': scale_x,
                'scale_y': scale_y
            }

        def map_screen_to_capture(screen_x, screen_y, capture_context):
            """
            功能:
                将屏幕绝对坐标映射为截图像素坐标
            参数:
                screen_x, screen_y: 屏幕绝对坐标
                capture_context: 映射上下文
            返回:
                元组, (img_x, img_y)
            """
            img_x = int(round((screen_x - capture_context['left']) * capture_context['scale_x']))
            img_y = int(round((screen_y - capture_context['top']) * capture_context['scale_y']))
            return img_x, img_y

        def create_picker_window():
            """
            功能:
                创建全屏透明取色窗口, 捕获鼠标点击
            参数:
                无
            返回:
                无
            """
            nonlocal picker_window, preview_frame

            # 先截取当前屏幕作为背景
            root.update()
            if table_window and table_window.winfo_exists():
                table_window.update()

            # 延迟一下确保窗口完全最小化
            root.after(100)
            try:
                screenshot = ImageGrab.grab(all_screens=True)
            except TypeError:
                screenshot = ImageGrab.grab()
            except Exception as e:
                messagebox.showerror("错误", f"屏幕截图失败: {str(e)}")
                cancel_color_picker(None)
                return

            capture_context = build_capture_context(screenshot)

            # 创建顶层窗口
            picker_window = tk.Toplevel()
            picker_window.overrideredirect(True)
            picker_window.attributes('-topmost', True)

            # 覆盖整个虚拟桌面(含多显示器和负坐标), 避免跨屏后光标失去十字形态
            try:
                virtual_left = int(capture_context['left'])
                virtual_top = int(capture_context['top'])
                virtual_width = int(capture_context['virtual_width'])
                virtual_height = int(capture_context['virtual_height'])
                picker_window.geometry(f"{virtual_width}x{virtual_height}{virtual_left:+d}{virtual_top:+d}")
            except Exception:
                # 退回全屏模式(主屏)
                picker_window.attributes('-fullscreen', True)

            # 设置窗口完全透明
            try:
                picker_window.attributes('-alpha', 0.01)
            except:
                pass

            picker_window.config(cursor='crosshair')

            # 创建提示标签
            hint_label = tk.Label(
                picker_window,
                text="点击屏幕任意位置取色, 按ESC取消",
                font=('Arial', 14, 'bold'),
                bg='yellow',
                fg='black',
                padx=10,
                pady=5
            )
            hint_label.place(relx=0.5, rely=0.05, anchor='center')

            # 创建预览窗口 - 使用独立的Toplevel确保可见
            preview_frame = tk.Toplevel()
            preview_frame.attributes('-topmost', True)
            preview_frame.overrideredirect(True)  # 无边框
            preview_frame.geometry('180x100+10+60')  # 位置和大小
            preview_frame.config(bg='white', relief='solid', borderwidth=3, cursor='crosshair')

            # 预览标题
            title_label = tk.Label(preview_frame, text="颜色预览", font=('Arial', 9, 'bold'), bg='white')
            title_label.pack(pady=2)

            # 颜色显示区域
            color_display = tk.Label(preview_frame, text="", bg='white', height=2)
            color_display.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)

            # 十六进制值标签
            hex_label = tk.Label(preview_frame, text="#000000", font=('Arial', 11, 'bold'), bg='white')
            hex_label.pack(pady=2)

            # 保存截图供取色使用
            picker_window.screenshot = screenshot
            picker_window.capture_context = capture_context

            # 绑定事件到主窗口
            picker_window.bind('<Button-1>', on_screen_click)
            picker_window.bind('<Motion>', lambda e: update_color_preview(e, color_display, hex_label, screenshot, capture_context))
            picker_window.bind('<Escape>', cancel_color_picker)

            # 绑定事件到预览窗口
            preview_frame.bind('<Button-1>', on_screen_click)
            preview_frame.bind('<Motion>', lambda e: update_color_preview(e, color_display, hex_label, screenshot, capture_context))
            preview_frame.bind('<Escape>', cancel_color_picker)

            # 保存预览窗口引用以便关闭
            picker_window.preview_window = preview_frame

            # 确保焦点在主窗口上以接收ESC键
            picker_window.lift()
            picker_window.focus_force()

        def update_color_preview(event, color_display, hex_label, screenshot, capture_context):
            """
            功能:
                实时更新颜色预览
            参数:
                event: 鼠标事件对象
                color_display: 颜色显示标签
                hex_label: 十六进制值标签
                screenshot: 预先截取的屏幕图像
                capture_context: 坐标映射上下文
            返回:
                无
            """
            try:
                # 读取鼠标真实屏幕坐标, 避免事件坐标在不同DPI下产生偏差
                if picker_window and picker_window.winfo_exists():
                    screen_x, screen_y = picker_window.winfo_pointerxy()
                else:
                    screen_x, screen_y = event.x_root, event.y_root
                x, y = map_screen_to_capture(screen_x, screen_y, capture_context)

                # 确保坐标在屏幕范围内
                width, height = screenshot.size
                if 0 <= x < width and 0 <= y < height:
                    rgb = screenshot.getpixel((x, y))
                    hex_color = '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])
                    color_display.config(bg=hex_color)
                    hex_label.config(text=hex_color)
            except Exception as e:
                pass

        def on_screen_click(event):
            """
            功能:
                处理屏幕点击事件, 获取点击位置的颜色
            参数:
                event: 鼠标事件对象, 包含点击坐标
            返回:
                无
            """
            nonlocal picker_window

            try:
                if not picker_window or not picker_window.winfo_exists():
                    return

                # 获取鼠标绝对坐标
                screen_x, screen_y = picker_window.winfo_pointerxy()

                # 从预先截取的屏幕图像中获取颜色
                screenshot = picker_window.screenshot
                capture_context = picker_window.capture_context
                x, y = map_screen_to_capture(screen_x, screen_y, capture_context)

                width, height = screenshot.size
                if not (0 <= x < width and 0 <= y < height):
                    raise ValueError("点击坐标超出屏幕截图范围")

                rgb_color = screenshot.getpixel((x, y))

                # 转换为十六进制格式
                hex_color = '#{:02x}{:02x}{:02x}'.format(rgb_color[0], rgb_color[1], rgb_color[2])

                # 关闭预览窗口
                if hasattr(picker_window, 'preview_window'):
                    picker_window.preview_window.destroy()

                # 关闭取色窗口
                picker_window.destroy()
                picker_window = None

                # 恢复主窗口
                root.deiconify()
                if table_window and table_window.winfo_exists():
                    table_window.deiconify()

                # 更新单元格颜色
                apply_picked_color(hex_color)

            except ImportError:
                messagebox.showerror("错误", "PIL库未安装, 请运行: pip install Pillow")
                cancel_color_picker(None)
            except OSError as e:
                messagebox.showerror("错误", f"屏幕截图失败: {str(e)}")
                cancel_color_picker(None)
            except Exception as e:
                messagebox.showerror("错误", f"取色失败: {str(e)}")
                cancel_color_picker(None)

        def apply_picked_color(hex_color):
            """
            功能:
                将取得的颜色应用到选中的单元格
            参数:
                hex_color: 字符串, 十六进制颜色代码, 如'#FF0000'
            返回:
                无
            """
            nonlocal picker_state

            row = picker_state['target_row']
            col = picker_state['target_col']

            # 更新单元格值
            table.set_cell_data(row, col, hex_color)

            # 更新单元格背景和前景色
            table.highlight_cells(
                row=row,
                column=col,
                bg=hex_color,
                fg=get_contrast_color(hex_color)
            )

            # 刷新图表
            update_plot()

            # 清空状态
            picker_state['target_row'] = None
            picker_state['target_col'] = None

        def cancel_color_picker(event):
            """
            功能:
                取消取色操作, 恢复窗口状态
            参数:
                event: 键盘事件对象(ESC键)
            返回:
                无
            """
            nonlocal picker_window, picker_state

            # 关闭预览窗口
            if picker_window and hasattr(picker_window, 'preview_window'):
                try:
                    picker_window.preview_window.destroy()
                except:
                    pass

            # 关闭取色窗口
            if picker_window:
                picker_window.destroy()
                picker_window = None

            # 恢复主窗口
            root.deiconify()
            if table_window and table_window.winfo_exists():
                table_window.deiconify()

            # 清空状态
            picker_state['target_row'] = None
            picker_state['target_col'] = None

        # 将右键菜单绑定到表格
        right_click_menu = RightClickMenu(table, start_color_picker)

            # 添加新列的函数
        def add_column():
            """
            功能:
                在表格末尾添加新列
            参数:
                无
            返回:
                无
            """
            # 获取当前列数
            current_cols = table.total_columns()

            # 在末尾插入新列
            table.insert_column(idx=current_cols, width=100)

            # 为所有行在新列填充空值
            total_rows = table.total_rows()
            for row_idx in range(total_rows):
                table.set_cell_data(row_idx, current_cols, '')

            # 更新表头
            new_header = f'E{current_cols - 2}'
            current_headers = list(table.headers())
            current_headers.append(new_header)
            table.headers(current_headers)

        def add_row():
            """
            功能:
                在表格末尾添加新行
            参数:
                无
            返回:
                无
            """
            num_columns = table.total_columns()
            new_row_values = ['#000000', '#000000', '#000000'] + [''] * (num_columns - 3)

            # 在末尾插入新行
            new_row_idx = table.total_rows()
            table.insert_row(row=new_row_values, idx=new_row_idx)

            # 设置前3列的背景颜色
            for col_idx in range(3):
                table.highlight_cells(
                    row=new_row_idx,
                    column=col_idx,
                    bg='#000000',
                    fg='#FFFFFF'
                )

            # 取消新添加行的选中状态
            table.deselect("all")

        def delete_column():
            """
            功能:
                删除表格的最后一列
            参数:
                无
            返回:
                无
            """
            total_cols = table.total_columns()

            # 确保至少保留3列
            if total_cols <= 3:
                print("无法删除: 必须至少保留3列")
                return

            # 删除最后一列
            table.delete_columns([total_cols - 1])

            # 更新表头
            new_headers = ['Curve Color', 'Marker Color', 'Text Color'] + \
                         [f'E{i+1}' for i in range(total_cols - 4)]
            table.headers(new_headers)

        def delete_row():
            """
            功能:
                删除表格的最后一行
            参数:
                无
            返回:
                无
            """
            total_rows = table.total_rows()

            # 检查是否有行可以删除
            if total_rows > 0:
                # 删除最后一行
                table.delete_rows([total_rows - 1])
            else:
                print("表格为空, 无法删除行")

        # 用于跟踪当前打开的文件路径
        current_file_path = [None]  # 使用列表以便在嵌套函数中修改

        # Function to save the table data
        def save_table_data():
            """
            功能:
                保存表格数据到文件, 支持json/xlsx/csv格式
            参数:
                无
            返回:
                无
            """
            # 收集表格数据
            data = []
            total_rows = table.total_rows()
            total_cols = table.total_columns()

            for row in range(total_rows):
                row_data = []
                for col in range(total_cols):
                    cell_value = table.get_cell_data(row, col)
                    row_data.append(cell_value)
                data.append(row_data)

            if not data:
                messagebox.showwarning("警告", "表格为空, 没有数据可保存")
                return

            # 如果有当前打开的文件, 提示是否保存到该文件或另存为
            if current_file_path[0]:
                response = messagebox.askyesnocancel(
                    "保存数据",
                    f"覆盖文件 {current_file_path[0]}?"
                )

                if response is None:  # User clicked cancel
                    print("保存操作已取消")
                    return
                elif response:  # User clicked yes, save to current file
                    file_path = current_file_path[0]
                else:  # User clicked no, save as
                    file_path = filedialog.asksaveasfilename(
                        title="另存为",
                        defaultextension=".json",
                        filetypes=[
                            ("JSON files", "*.json"),
                            ("Excel files", "*.xlsx"),
                            ("CSV files", "*.csv"),
                            ("All files", "*.*")
                        ],
                        initialfile=os.path.basename(current_file_path[0])
                    )
                    if not file_path:
                        print("保存操作已取消")
                        return
            else:
                # 没有当前文件, 显示保存对话框
                file_path = filedialog.asksaveasfilename(
                    title="保存数据",
                    defaultextension=".json",
                    filetypes=[
                        ("JSON files", "*.json"),
                        ("Excel files", "*.xlsx"),
                        ("CSV files", "*.csv"),
                        ("All files", "*.*")
                    ],
                    initialfile="table_data.json"
                )

                if not file_path:
                    print("保存操作已取消")
                    return

            # 根据文件扩展名选择保存方法
            try:
                file_ext = os.path.splitext(file_path)[1].lower()

                if file_ext == '.json':
                    # 保存为JSON格式
                    with open(file_path, 'w', encoding='utf-8') as file:
                        json.dump(data, file, ensure_ascii=False, indent=2)

                elif file_ext == '.xlsx':
                    # 保存为Excel格式
                    headers = list(table.headers())
                    df = pd.DataFrame(data, columns=headers)

                    # 检查文件是否已存在
                    file_exists = os.path.exists(file_path)

                    if file_exists:
                        # 文件已存在, 读入内存操作避免锁定磁盘文件
                        with open(file_path, 'rb') as f:
                            file_data = BytesIO(f.read())

                        wb = load_workbook(file_data)
                        ws = wb.active

                        # 清除现有数据(保留格式)
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                            for cell in row:
                                cell.value = None

                        # 写入新数据
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                            for c_idx, value in enumerate(row, 1):
                                ws.cell(row=r_idx, column=c_idx, value=value)

                        # 保存到内存再写回磁盘
                        output = BytesIO()
                        wb.save(output)
                        wb.close()
                        del wb
                        gc.collect()

                        with open(file_path, 'wb') as f:
                            f.write(output.getvalue())
                    else:
                        # 文件不存在, 在内存中创建后写入磁盘
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='Sheet1')

                            # 设置列宽
                            ws = writer.sheets['Sheet1']
                            for column in ws.columns:
                                max_length = 18
                                column_letter = column[0].column_letter
                                ws.column_dimensions[column_letter].width = max_length

                        with open(file_path, 'wb') as f:
                            f.write(output.getvalue())

                elif file_ext == '.csv':
                    # 保存为CSV格式
                    headers = list(table.headers())
                    df = pd.DataFrame(data, columns=headers)
                    df.to_csv(file_path, index=False, encoding='utf-8-sig')

                else:
                    # 默认保存为JSON格式
                    with open(file_path, 'w', encoding='utf-8') as file:
                        json.dump(data, file, ensure_ascii=False, indent=2)

                # 更新当前文件路径
                current_file_path[0] = file_path
                print(f"表格数据已保存到: {file_path}")
                messagebox.showinfo("成功", f"数据已保存到:\n{file_path}")

            except Exception as e:
                print(f"保存失败: {str(e)}")
                messagebox.showerror("错误", f"保存失败:\n{str(e)}")

        # Function to load the table data
        def load_table_data():
            """
            功能:
                从文件加载表格数据, 支持json/xlsx/csv格式
            参数:
                无
            返回:
                无
            """
            # Show file selection dialog
            file_path = filedialog.askopenfilename(
                title="加载数据",
                filetypes=[
                    ("All supported files", "*.json;*.xlsx;*.csv"),
                    ("JSON files", "*.json"),
                    ("Excel files", "*.xlsx"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )

            if not file_path:
                print("加载操作已取消")
                return

            try:
                # 根据文件扩展名选择加载方法
                file_ext = os.path.splitext(file_path)[1].lower()
                data = []

                if file_ext == '.json':
                    # 从JSON文件加载
                    with open(file_path, 'r', encoding='utf-8') as file:
                        data = json.load(file)

                elif file_ext == '.xlsx':
                    # 从Excel文件加载, 读入内存避免锁定磁盘文件
                    with open(file_path, 'rb') as f:
                        file_data = BytesIO(f.read())
                    df = pd.read_excel(file_data, engine='openpyxl')
                    # 将NaN值替换为空字符串
                    df = df.fillna('')
                    data = df.values.tolist()

                elif file_ext == '.csv':
                    # 从CSV文件加载
                    df = pd.read_csv(file_path, encoding='utf-8-sig')
                    # 将NaN值替换为空字符串
                    df = df.fillna('')
                    data = df.values.tolist()

                else:
                    messagebox.showerror("错误", f"不支持的文件格式: {file_ext}")
                    return

                if not data:
                    messagebox.showwarning("警告", "文件中没有数据")
                    return

                # 规范化数据结构，确保列数一致且至少包含前三列颜色列
                normalized_data = []
                max_cols = 0
                for row_data in data:
                    if isinstance(row_data, (list, tuple)):
                        row_values = list(row_data)
                    else:
                        row_values = [row_data]
                    normalized_data.append(row_values)
                    max_cols = max(max_cols, len(row_values))

                num_cols = max(3, max_cols)
                for row_values in normalized_data:
                    if len(row_values) < num_cols:
                        row_values.extend([''] * (num_cols - len(row_values)))
                    for color_col in range(3):
                        if row_values[color_col] in ('', None):
                            row_values[color_col] = '#000000'

                # 先设置完整数据，再更新表头和列宽，避免列扩展过程中的越界
                table.set_sheet_data(normalized_data)
                headers = ['Curve Color', 'Marker Color', 'Text Color'] + [f'E{i+1}' for i in range(num_cols - 3)]
                table.headers(headers)
                for col_idx in range(num_cols):
                    table.column_width(column=col_idx, width=100)

                # 设置表格所有显示内容居中对齐（数据区、表头、行号）
                table.table_align(align="center")
                table.header_align(align="center")
                table.row_index_align(align="center")

                # 更新前3列的背景颜色
                for row in range(len(normalized_data)):
                    for col in range(3):
                        color_value = table.get_cell_data(row, col)
                        if color_value and isinstance(color_value, str) and color_value.startswith('#'):
                            table.highlight_cells(
                                row=row,
                                column=col,
                                bg=color_value,
                                fg=get_contrast_color(color_value)
                            )

                # 取消所有选中状态, 避免加载后所有内容都是蓝色选中状态
                table.deselect("all")

                show_all()

                # 更新当前文件路径
                current_file_path[0] = file_path
                print(f"表格数据已从以下位置加载: {file_path}")
                messagebox.showinfo("成功", f"数据已从以下位置加载:\n{file_path}")

            except Exception as e:
                print(f"加载失败: {str(e)}")
                messagebox.showerror("错误", f"加载失败:\n{str(e)}")
        
        add_row_button = tk.Button(buttons_frame, text="Add Row", command=add_row)
        add_column_button = tk.Button(buttons_frame, text="Add Column", command=add_column)
        delete_row_button = tk.Button(buttons_frame, text="Delete Row", command=delete_row)
        delete_column_button = tk.Button(buttons_frame, text="Delete Column", command=delete_column)

        save_button = tk.Button(buttons_frame, text="Save Data", command=save_table_data)
        load_button = tk.Button(buttons_frame, text="Load Data", command=load_table_data)
        pick_color_button = tk.Button(buttons_frame, text="Pick Color", command=start_color_picker)


        add_row_button.pack(side=tk.LEFT, padx=5, pady=2)
        add_column_button.pack(side=tk.LEFT, padx=5, pady=2)
        delete_row_button.pack(side=tk.LEFT, padx=5, pady=2)
        delete_column_button.pack(side=tk.LEFT, padx=5, pady=2)
        save_button.pack(side=tk.LEFT, padx=5, pady=2)
        load_button.pack(side=tk.LEFT, padx=5, pady=2)
        pick_color_button.pack(side=tk.LEFT, padx=5, pady=2)

        return table_window, table

    global table, table_window

    table_window, table = create_table_window()
    # 存储所有动态创建的滑条
    slider_frames = {}

    def add_slider_with_entry(label, variable, from_, to, resolution, row, length=100):

        #检查是否已有相同 label 的滑条，若存在则销毁
        if label in slider_frames:
            slider_frames[label].destroy()
            del slider_frames[label]

        # 创建一个新的 Frame 用于包装每一组 slider 和 entry
        row_frame = tk.Frame(control_frame,height=10)
        row_frame.pack(fill='y', padx=5, pady=2,expand=True)

        #存储滑条引用
        slider_frames[label] = row_frame 
        # 设置标签宽度，确保它们具有一致的宽度
        label_widget = ttk.Label(row_frame, text=label, width=12)  # 设定固定宽度
        label_widget.pack(side=tk.LEFT, padx=5, anchor='s')

        # 确定分辨率的小数位数
        precision = len(str(resolution).split('.')[-1]) if '.' in str(resolution) else 0

        # 创建一个减少按钮
        def decrease_value():
            new_value = round(variable.get() - resolution, precision)
            if new_value >= from_:
                variable.set(new_value)

        def decrease_hold():
            decrease_value()
            global hold_decrease
            hold_decrease = row_frame.after(50, decrease_hold)  # 50ms 后继续调用

        def stop_decrease():
            global hold_decrease
            row_frame.after_cancel(hold_decrease)

        decrease_button = ttk.Button(row_frame, text='-', width=2)
        decrease_button.pack(side=tk.LEFT, padx=2, anchor='s')
        decrease_button.bind("<ButtonPress-1>", lambda e: decrease_hold())
        decrease_button.bind("<ButtonRelease-1>", lambda e: stop_decrease())

        # 创建滑动条并添加到内部 Frame 使用 pack 布局
        slider = tk.Scale(row_frame, from_=from_, to=to, variable=variable, orient=tk.HORIZONTAL,
                        resolution=resolution, length=length, showvalue=False)
        slider.pack(side=tk.LEFT, padx=5, pady=2, anchor='s')  # 使用 pack 来放置滑动条

        # 创建一个增加按钮
        def increase_value():
            new_value = round(variable.get() + resolution, precision)
            if new_value <= to:
                variable.set(new_value)

        def increase_hold():
            increase_value()
            global hold_increase
            hold_increase = row_frame.after(50, increase_hold)  # 50ms 后继续调用

        def stop_increase():
            global hold_increase
            row_frame.after_cancel(hold_increase)

        increase_button = ttk.Button(row_frame, text='+', width=2)
        increase_button.pack(side=tk.LEFT, padx=2, anchor='s')
        increase_button.bind("<ButtonPress-1>", lambda e: increase_hold())
        increase_button.bind("<ButtonRelease-1>", lambda e: stop_increase())

        # 创建输入框并添加到当前行
        entry = ttk.Entry(row_frame, textvariable=variable, width=8)
        entry.pack(side=tk.LEFT, padx=5, anchor='s')  # 保持输入框使用 pack 布局

        # 为输入框绑定事件，按回车键时更新图形
        entry.bind("<Return>", update_plot)

        # 为滑块和输入框添加事件，使得值变化时都能更新图形
        variable.trace_add("write", update_plot)
        
        
    # 创建主窗口和控制面板
    control_frame = tk.Frame(root)
    control_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5, pady=5)

    # 配置控件所在列的权重
    control_frame.columnconfigure(1, weight=1)

    # 定义可调节的变量
    adjustment_factor_var = tk.DoubleVar(value=0.0)
    width_factor_var = tk.DoubleVar(value=0.3)
    adjustment_factor_2_var = tk.DoubleVar(value=0.0)
    scale_factor_var = tk.DoubleVar(value=10.0)
    scale_factor_x_var = tk.DoubleVar(value=1.0)
    scale_factor_y_var = tk.DoubleVar(value=1.0)
    bond_length_var = tk.DoubleVar(value=20.0)
    radius_var = tk.DoubleVar(value=3.0)
    font_size_var = tk.DoubleVar(value=10.0)
    text_space_var = tk.DoubleVar(value=3.0)
    curve_width_var = tk.DoubleVar(value=0.6)
    bond_width_var = tk.DoubleVar(value=2.0)

    # 添加滑块和输入框
    add_slider_with_entry("Scale Factor", scale_factor_var, 0, 20.0, 0.1, 3)
    add_slider_with_entry("Scale Factor x", scale_factor_x_var, 0.0, 5, 0.02, 3)
    add_slider_with_entry("Scale Factor y", scale_factor_y_var, 0.0, 10.0, 0.1, 3)
    add_slider_with_entry("Base Curve", width_factor_var, 0.0, 1.0, 0.01, 0)
    add_slider_with_entry("Disp Factor", adjustment_factor_var, 0.0, 2, 0.01, 1)
    add_slider_with_entry("Curve Factor", adjustment_factor_2_var, 0.0, 0.02, 0.0001, 2)
    add_slider_with_entry("Curve Width", curve_width_var, 0.0, 5.0, 0.1, 2)
    add_slider_with_entry("Font Size", font_size_var, 0.0, 100.0, 1, 6)
    add_slider_with_entry("Text Space", text_space_var, 0.0, 50, 1, 6)

    add_slider_with_entry("Bond Length", bond_length_var, 0.0, 40.0, 0.5, 4)
    add_slider_with_entry("Bond Width", bond_width_var, 0.0, 10.0, 0.1, 2)
    add_slider_with_entry("Circle Radius", radius_var, 0.0, 10.0, 0.1, 5)

    # 创建形状选择框
    shape_frame = tk.Frame(control_frame)
    shape_frame.pack(padx=5, pady=2, anchor='w',fill='y',expand=True)

    tk.Label(shape_frame, text="Shape:").pack(side=tk.LEFT, padx=5)
    shape_type_var = tk.StringVar(value="line")
    tk.Radiobutton(shape_frame, text="Line", variable=shape_type_var, value="line", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(shape_frame, text="Circle", variable=shape_type_var, value="circle", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(shape_frame, text="None", variable=shape_type_var, value="None", command=update_plot).pack(side=tk.LEFT, padx=10)

    # 创建连接选择框
    connection_frame = tk.Frame(control_frame)
    connection_frame.pack(padx=5, pady=2, anchor='w',fill='y',expand=True)

    tk.Label(connection_frame, text="Connection:").pack(side=tk.LEFT, padx=5)
    connect_type_var = tk.StringVar(value="center")
    tk.Radiobutton(connection_frame, text="Center", variable=connect_type_var, value="center", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(connection_frame, text="Side", variable=connect_type_var, value="side", command=update_plot).pack(side=tk.LEFT, padx=5)

    # 创建线性选择框
    line_type_frame = tk.Frame(control_frame)
    line_type_frame.pack(padx=5, pady=2, anchor='w',fill='y',expand=True)

    tk.Label(line_type_frame, text="Line Type:").pack(side=tk.LEFT, padx=5)
    line_type_var = tk.StringVar(value="Solid")
    tk.Radiobutton(line_type_frame, text="Solid", variable=line_type_var, value="Solid", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(line_type_frame, text="Dashed", variable=line_type_var, value="Dashed", command=update_plot).pack(side=tk.LEFT, padx=5)

    # 创建能量和标签的连接形式框
    target_layout_frame = tk.Frame(control_frame)
    target_layout_frame .pack(padx=5, pady=2, anchor='w',fill='y',expand=True)

    tk.Label(target_layout_frame , text="Label-Energy Layout").pack(side=tk.LEFT, padx=5)
    target_layout_var = tk.StringVar(value="sperate")
    tk.Radiobutton(target_layout_frame, text="Sperate", variable= target_layout_var, value="sperate", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(target_layout_frame, text="Combine", variable= target_layout_var, value="combine", command=update_plot).pack(side=tk.LEFT, padx=5)

    # 创建标签位置选择框
    target_location_frame = tk.Frame(control_frame)
    target_location_frame .pack(padx=5, pady=2, anchor='w',fill='y',expand=True)

    tk.Label(target_location_frame , text="Target Location").pack(side=tk.LEFT, padx=5)
    target_location_var = tk.StringVar(value="c")
    tk.Radiobutton(target_location_frame, text="C", variable= target_location_var, value="c", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(target_location_frame, text="W", variable= target_location_var, value="w", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(target_location_frame, text="S", variable= target_location_var, value="s", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(target_location_frame, text="A", variable= target_location_var, value="a", command=update_plot).pack(side=tk.LEFT, padx=5)
    tk.Radiobutton(target_location_frame, text="D", variable= target_location_var, value="d", command=update_plot).pack(side=tk.LEFT, padx=5)

    # 使用一个子Frame排列选项框
    options_frame = tk.Frame(control_frame)
    options_frame.pack(fill=tk.X, expand=True)

    show_marker_var = tk.BooleanVar(value=False)
    checkbutton = tk.Checkbutton(options_frame, text="Show Marker", variable=show_marker_var, command=update_plot)
    checkbutton.pack(side=tk.LEFT,padx=5, pady=2, anchor='w',fill='y')

    grid_var = tk.BooleanVar(value=True)
    grid_checkbutton = tk.Checkbutton(options_frame, text="Show Grid", variable=grid_var, command=update_plot)
    grid_checkbutton.pack(side=tk.LEFT,padx=5, pady=2, anchor='w',fill='y')

    # 创建新的一行用于Show Numbers和Show Targets
    options_frame_2 = tk.Frame(control_frame)
    options_frame_2.pack(fill=tk.X, expand=True)

    show_numbers_var = tk.BooleanVar(value=True)
    numbers_checkbutton = tk.Checkbutton(options_frame_2, text="Show Numbers", variable=show_numbers_var, command=update_plot)
    numbers_checkbutton.pack(side=tk.LEFT,padx=5, pady=2, anchor='w',fill='y')

    show_targets_var = tk.BooleanVar(value=True)
    targets_checkbutton = tk.Checkbutton(options_frame_2, text="Show Targets", variable=show_targets_var, command=update_plot)
    targets_checkbutton.pack(side=tk.LEFT,padx=5, pady=2, anchor='w',fill='y')


    def hex_to_rgb(hex_color):
        """Convert hexadecimal color code to RGB tuple"""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) / 255.0 for i in (0, 2, 4))

    def add_colors_to_colortable(color_list):
        """
        功能:
            将颜色列表添加到颜色表, 保留原始颜色
        参数:
            color_list: 列表, 十六进制颜色代码列表
        返回:
            字符串, XML格式的颜色表
        """
        # 创建根元素<colortable>
        colortable = ET.Element("colortable")

        # 保留原始颜色
        original_colors = [
            "#FFFFFF",  # 白色
            "#000000",  # 黑色
            "#FF0000",  # 红色
            "#FFFF00",  # 黄色
            "#00FF00",  # 绿色
            "#00FFFF",  # 青色
            "#0000FF",  # 蓝色
            "#FF00FF"   # 品红色
        ]

        # 添加原始颜色
        for color in original_colors:
            r, g, b = hex_to_rgb(color)
            color_element = ET.SubElement(colortable, "color", r=str(r), g=str(g), b=str(b))

        # 从颜色列表添加额外的颜色
        for color in color_list:
            r, g, b = hex_to_rgb(color)
            color_element = ET.SubElement(colortable, "color", r=str(r), g=str(g), b=str(b))

        # 创建一个ElementTree对象并生成XML字符串
        xml_str = ET.tostring(colortable, encoding="unicode", method="xml")

        return xml_str
    
    def export_cdxml(*args):
        global already_draw_target
        already_draw_target = []
        nonlocal loaded_cdxml_header, loaded_font_xml, loaded_font_id
        try:
            # 获取用户输入参数
            adjustment_factor = float(adjustment_factor_var.get())
            width_factor = float(width_factor_var.get())
            adjustment_factor_2 = float(adjustment_factor_2_var.get())
            scale_factor = float(scale_factor_var.get())
            scale_factor_x = float(scale_factor_x_var.get())
            scale_factor_y = float(scale_factor_y_var.get())
            bond_length = float(bond_length_var.get())
            radius = float(radius_var.get())
            connect_type = connect_type_var.get()
            shape_type = shape_type_var.get()
            line_type = line_type_var.get()
            font_size = font_size_var.get()
            text_space = text_space_var.get()
            target_layout = target_layout_var.get()
            target_location = target_location_var.get()
            curve_width = curve_width_var.get()
            bond_width = bond_width_var.get()

            # 确定使用的字体ID
            current_font_id = int(loaded_font_id) if loaded_font_id is not None else 3
            target_layout = target_layout_var.get()
            target_location = target_location_var.get()
            curve_width = curve_width_var.get()
            bond_width = bond_width_var.get()


            connect_xml = ''
            graph_xml = ''
            color_list  = [] # 颜色目标
            global page_high
            global page_width
            page_high = 1
            page_width = 1
            y_limit = 300

            # 从表格读取数据(使用tksheet的正确方法)
            total_rows = table.total_rows()

            for row in range(total_rows):
                values = []
                for col in range(table.total_columns()):
                    values.append(table.get_cell_data(row, col))

                # 跳过结构不完整的临时行
                if len(values) < 3:
                    continue

                for i in values[:3]: # 颜色目标
                    color_list.append(i)

            color_list = list(set(color_list)) # 移除重复颜色
            color_xml = add_colors_to_colortable(color_list)

            # 设置y限制
            total_y= []
            for row in range(total_rows):
                values = []
                for col in range(table.total_columns()):
                    values.append(table.get_cell_data(row, col))

                # 跳过结构不完整的临时行
                if len(values) < 3:
                    continue

                for index, value in enumerate(values[3:]):
                    value = str(value)
                    if value.strip():  # 如果值不为空
                        try:
                            value = parse_data(value)
                            total_y.append(float(value[0]))

                        except ValueError:
                            print(f"Invalid value at row {row}, column {index + 2}. Please correct.")
                            return

            scaled_y_points = [y * scale_factor*scale_factor_y / 2 for y in total_y]

            for i in scaled_y_points:
                if i > y_limit :
                    y_limit = i+100

            # 生成曲线和目标数据
            for row in range(total_rows):
                values = []
                for col in range(table.total_columns()):
                    values.append(table.get_cell_data(row, col))

                # 跳过结构不完整的临时行
                if len(values) < 3:
                    continue

                original_y = []
                target = []
                location = []
                original_x = []
                for index, value in enumerate(values[3:]):
                    value = str(value)
                    if value.strip():  # 如果值不为空
                        try:
                            value = parse_data(value)
                            original_y.append(float(value[0]))
                            target.append(value[1])
                            location.append(value[2])
                            original_x.append(index + 1)

                        except ValueError:
                            print(f"Invalid value at row {row}, column {index + 2}. Please correct.")
                            return

                curve_color = values[0]
                marker_color = values[1]
                text_color = values[2]


                # 调用贝塞尔曲线计算函数
                x_adjusted, y_points, bezier_curves = get_bezier_curve_points_flat(
                   original_x, original_y, adjustment_factor, width_factor, adjustment_factor_2
                )

                # 缩放贝塞尔曲线
                if  x_adjusted[0] == 'no_data': # 跳过空行
                    continue
                # 缩放贝塞尔曲线
                for curve in bezier_curves:
                    X = curve['X']
                    Y = curve['Y']
                    for i in range(len(X)):
                        X[i] = X[i] * scale_factor *scale_factor_x* 10 + 50
                        Y[i] = y_limit - Y[i] * scale_factor*scale_factor_y / 2

                # 缩放调整后的坐标和势能值
                scaled_x_adjusted = [x * scale_factor*scale_factor_x * 10 + 50 for x in x_adjusted]
                scaled_y_points = [y_limit - y * scale_factor*scale_factor_y / 2 for y in y_points]

                # 自动调整页面宽度和高度
                for x in  scaled_x_adjusted:
                    ad_page_width = math.ceil(abs(x/510))
                    if ad_page_width > page_width:
                        page_width = ad_page_width
                for y in scaled_y_points:
                    ad_page_high = math.ceil(abs(y/710))
                    if ad_page_high > page_high:
                        page_high = ad_page_high

                curve_color_index = color_list.index(curve_color) + 10
                marker_color_index = color_list.index(marker_color) + 10
                text_color_index = color_list.index(text_color) + 10

                text_base_movement = round(font_size*0.3241+0.1236,2)


                if shape_type ==  "line":
                    if width_factor == 0:
                        curves_xml = draw_line(scaled_x_adjusted,scaled_y_points,line_type,curve_width,bond_length,curve_color_index,connect_type,original_x)
                    else:
                        curves_xml = draw_curve(bezier_curves, connect_type, bond_length,curve_color_index,line_type,original_x,curve_width)
                    graph_xml  += "\n".join(curves_xml)
                    rectangle_xml = generate_rectangle_xml(scaled_x_adjusted, scaled_y_points, bond_length, marker_color_index, connect_type,original_x,bond_width)
                    graph_xml  += "\n".join(rectangle_xml)
                    text_xml = generate_text_cdxml(scaled_x_adjusted, scaled_y_points, original_y,target, location,target_layout,target_location,text_space,text_base_movement,bond_length,bond_width/2,text_color_index, connect_type, original_x,font_size,font_type=current_font_id, Z_value=70, show_numbers=show_numbers_var.get(), show_targets=show_targets_var.get())
                    graph_xml  += "\n".join(text_xml)

                elif shape_type == "circle":
                    if width_factor == 0:
                        curves_xml = draw_line(scaled_x_adjusted,scaled_y_points,line_type,curve_width,radius,curve_color_index,connect_type,original_x)
                    else:
                        curves_xml = draw_curve(bezier_curves, connect_type, radius,curve_color_index,line_type,original_x,curve_width)
                    graph_xml  += "\n".join(curves_xml)
                    circles_xml = generate_circle_xml(scaled_x_adjusted, scaled_y_points, radius, marker_color_index, connect_type,original_x)
                    graph_xml  += "\n".join(circles_xml)
                    text_xml = generate_text_cdxml(scaled_x_adjusted, scaled_y_points, original_y,target, location,target_layout,target_location,text_space,text_base_movement,radius,radius,text_color_index, connect_type, original_x,font_size,font_type=current_font_id, Z_value=70, show_numbers=show_numbers_var.get(), show_targets=show_targets_var.get())
                    graph_xml  += "\n".join(text_xml)
                elif shape_type == "None":
                    if width_factor == 0:
                        curves_xml = draw_line(scaled_x_adjusted,scaled_y_points,line_type,curve_width,0,curve_color_index,connect_type,original_x)
                    else:
                        curves_xml = draw_curve(bezier_curves, connect_type,0,curve_color_index,line_type,original_x,curve_width)
                    graph_xml  += "\n".join(curves_xml)
                    text_xml = generate_text_cdxml(scaled_x_adjusted, scaled_y_points, original_y,target, location,target_layout,target_location,text_space,text_base_movement,0,0,text_color_index, connect_type, original_x,font_size,font_type=current_font_id, Z_value=70, show_numbers=show_numbers_var.get(), show_targets=show_targets_var.get())
                    graph_xml  += "\n".join(text_xml)

            page_xml = f'''<page
                id="66"
                BoundingBox="0 0 1620 719.75"
                HeaderPosition="36"
                FooterPosition="36"
                PrintTrimMarks="yes"
                HeightPages="{page_high}"
                WidthPages="{page_width}"
                >'''

            connect_xml  += color_xml  # 添加颜色
            # 使用读取的font_xml, 如果没有读取则使用默认值
            current_font_xml = loaded_font_xml if loaded_font_xml is not None else font_xml
            connect_xml  += current_font_xml  # 添加字体
            connect_xml  += page_xml  # 添加页面
            connect_xml  += graph_xml  # 添加图形

            # 将所有部分组合成单个CDXML字符串
            # 使用读取的cdxml_header, 如果没有读取则使用默认值
            current_cdxml_header = loaded_cdxml_header if loaded_cdxml_header is not None else cdxml_header
            cdxml_string = current_cdxml_header + connect_xml + cdxml_footer

            # 保存CDXML文件
            save_cdxml_file(cdxml_string)

        except ValueError:
            print("Input error, please enter valid numeric values.")

    def load_cdxml_style():
        """
        功能:
            从CDXML文件读取风格信息(cdxml_header和font_xml)
        参数:
            无
        返回:
            无
        """
        nonlocal loaded_cdxml_header, loaded_font_xml, loaded_font_id, loaded_font_name

        # 显示文件选择对话框
        file_path = filedialog.askopenfilename(
            filetypes=[("CDXML files", "*.cdxml"), ("All files", "*.*")],
            title="选择CDXML文件以读取风格信息"
        )

        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                # 解析XML
                root = ET.fromstring(content)

                # 提取cdxml_header (CDXML标签的属性)
                cdxml_attrs = root.attrib
                header_lines = ['<?xml version="1.0" encoding="UTF-8" ?>']
                header_lines.append('<!DOCTYPE CDXML SYSTEM "http://www.cambridgesoft.com/xml/cdxml.dtd" >')
                header_lines.append('<CDXML')

                for key, value in cdxml_attrs.items():
                    # 转义特殊字符
                    value_escaped = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&apos;')
                    header_lines.append(f' {key}="{value_escaped}"')

                header_lines.append('>')
                loaded_cdxml_header = '\n'.join(header_lines)

                # 提取font_xml (fonttable标签)
                fonttable = root.find('fonttable')
                if fonttable is not None:
                    # 将fonttable转换为字符串
                    loaded_font_xml = ET.tostring(fonttable, encoding='unicode', method='xml')

                    # 提取第一个字体的ID和名称 (通常用于文本标签)
                    fonts = fonttable.findall('font')
                    if fonts:
                        # 优先使用非Symbol字体
                        for font in fonts:
                            font_name = font.get('name', '')
                            if font_name.lower() != 'symbol':
                                loaded_font_id = font.get('id')
                                loaded_font_name = font_name
                                break
                        # 如果没有找到非Symbol字体, 使用第一个字体
                        if loaded_font_id is None:
                            loaded_font_id = fonts[0].get('id')
                            loaded_font_name = fonts[0].get('name', 'Arial')
                else:
                    # 如果没有找到fonttable, 使用默认值
                    loaded_font_xml = '''<fonttable>
<font id="3" charset="iso-8859-1" name="Arial"/>
</fonttable>'''
                    loaded_font_id = '3'
                    loaded_font_name = 'Arial'

                # 刷新画布以应用新字体
                update_plot()
                messagebox.showinfo("成功", f"已从文件读取CDXML风格信息: {file_path}\n字体ID: {loaded_font_id}\n字体名称: {loaded_font_name}")

            except Exception as e:
                messagebox.showerror("错误", f"读取CDXML文件失败: {str(e)}")

    def save_config():
        """
        功能:
            将当前所有配置参数保存到JSON文件
        参数:
            无
        返回:
            无
        """
        nonlocal loaded_cdxml_header, loaded_font_xml, loaded_font_id, loaded_font_name
        config = {
            'adjustment_factor': adjustment_factor_var.get(),
            'width_factor': width_factor_var.get(),
            'adjustment_factor_2': adjustment_factor_2_var.get(),
            'scale_factor': scale_factor_var.get(),
            'scale_factor_x': scale_factor_x_var.get(),
            'scale_factor_y': scale_factor_y_var.get(),
            'bond_length': bond_length_var.get(),
            'radius': radius_var.get(),
            'font_size': font_size_var.get(),
            'text_space': text_space_var.get(),
            'curve_width': curve_width_var.get(),
            'bond_width': bond_width_var.get(),
            'shape_type': shape_type_var.get(),
            'connect_type': connect_type_var.get(),
            'line_type': line_type_var.get(),
            'target_layout': target_layout_var.get(),
            'target_location': target_location_var.get(),
            'show_marker': show_marker_var.get(),
            'grid': grid_var.get(),
            'cdxml_header': loaded_cdxml_header,
            'font_xml': loaded_font_xml,
            'font_id': loaded_font_id,
            'font_name': loaded_font_name
        }

        # 显示文件选择对话框
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Configuration"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=4, ensure_ascii=False)
                messagebox.showinfo("Success", f"Configuration saved to: {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save configuration: {str(e)}")

    def load_config():
        """
        功能:
            从JSON文件加载配置参数并应用到当前界面
        参数:
            无
        返回:
            无
        """
        nonlocal loaded_cdxml_header, loaded_font_xml, loaded_font_id, loaded_font_name
        # 显示文件选择对话框
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Configuration"
        )

        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                # 将配置应用到变量
                if 'adjustment_factor' in config:
                    adjustment_factor_var.set(config['adjustment_factor'])
                if 'width_factor' in config:
                    width_factor_var.set(config['width_factor'])
                if 'adjustment_factor_2' in config:
                    adjustment_factor_2_var.set(config['adjustment_factor_2'])
                if 'scale_factor' in config:
                    scale_factor_var.set(config['scale_factor'])
                if 'scale_factor_x' in config:
                    scale_factor_x_var.set(config['scale_factor_x'])
                if 'scale_factor_y' in config:
                    scale_factor_y_var.set(config['scale_factor_y'])
                if 'bond_length' in config:
                    bond_length_var.set(config['bond_length'])
                if 'radius' in config:
                    radius_var.set(config['radius'])
                if 'font_size' in config:
                    font_size_var.set(config['font_size'])
                if 'text_space' in config:
                    text_space_var.set(config['text_space'])
                if 'curve_width' in config:
                    curve_width_var.set(config['curve_width'])
                if 'bond_width' in config:
                    bond_width_var.set(config['bond_width'])
                if 'shape_type' in config:
                    shape_type_var.set(config['shape_type'])
                if 'connect_type' in config:
                    connect_type_var.set(config['connect_type'])
                if 'line_type' in config:
                    line_type_var.set(config['line_type'])
                if 'target_layout' in config:
                    target_layout_var.set(config['target_layout'])
                if 'target_location' in config:
                    target_location_var.set(config['target_location'])
                if 'show_marker' in config:
                    show_marker_var.set(config['show_marker'])
                if 'grid' in config:
                    grid_var.set(config['grid'])

                # 加载CDXML风格信息
                if 'cdxml_header' in config:
                    loaded_cdxml_header = config['cdxml_header']
                if 'font_xml' in config:
                    loaded_font_xml = config['font_xml']
                if 'font_id' in config:
                    loaded_font_id = config['font_id']
                if 'font_name' in config:
                    loaded_font_name = config['font_name']

                # 加载配置后更新绘图
                update_plot()
                messagebox.showinfo("Success", f"Configuration loaded from {file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load configuration: {str(e)}")

    def re_create_table_window():
        """
        功能:
            重新创建或显示表格窗口
        参数:
            无
        返回:
            无
        """
        global table, table_window

        # 检查表格窗口是否存在并且可见
        if table_window is None or not table_window.winfo_exists():
            # 如果没有窗口或者窗口已经被关闭, 创建新的窗口
            table_window, table = create_table_window()
        else:
            print("表格窗口已经打开")

    # 创建第一行按钮frame
    buttons_frame_1 = tk.Frame(control_frame)
    buttons_frame_1.pack(fill=tk.X, expand=True, pady=2)

    export_button = tk.Button(buttons_frame_1, text="Show All", command=show_all)
    export_button.pack(side=tk.LEFT, padx=5, pady=2)

    export_button = tk.Button(buttons_frame_1, text="Open Table", command=re_create_table_window)
    export_button.pack(side=tk.LEFT, padx=5, pady=2)

    # 添加保存配置按钮
    save_config_button = tk.Button(buttons_frame_1, text="Save Config", command=save_config)
    save_config_button.pack(side=tk.LEFT, padx=5, pady=2)

    # 添加加载配置按钮
    load_config_button = tk.Button(buttons_frame_1, text="Load Config", command=load_config)
    load_config_button.pack(side=tk.LEFT, padx=5, pady=2)

    # 创建新的一行用于Export CDXML按钮
    export_frame = tk.Frame(control_frame)
    export_frame.pack(fill=tk.X, expand=True, pady=2)

    # 添加读取CDXML风格按钮
    load_style_button = tk.Button(export_frame, text="Load CDXML Style", command=load_cdxml_style)
    load_style_button.pack(side=tk.LEFT, padx=5, pady=2)

    export_cdxml_button = tk.Button(export_frame, text="Export CDXML", command=export_cdxml)
    export_cdxml_button.pack(side=tk.LEFT, padx=5, pady=2)

    root.mainloop()

if __name__ == "__main__":
    interactive_bezier_curve()


